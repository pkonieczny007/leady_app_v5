# -*- coding: utf-8 -*-
"""
Scenariusze akceptacyjne — przejście przez to, o co klient poprosił, przez API aplikacji.

Uruchomienie:  python test_scenariusze.py
Test działa na WŁASNEJ, tymczasowej bazie (nie rusza `data/leady_v3.db`).

Każdy scenariusz jest nazwany językiem klienta, a nie technicznym, żeby dało się
je przeklikać razem z nim i zapytać „czy o to chodziło?".
"""
import datetime as dt
import json
import os
import shutil
import sys
import tempfile

# konsola Windows domyślnie cp1250 — bez tego polskie znaki wysypują wydruk
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

# własna baza na czas testu — zanim zaimportujemy moduły aplikacji
TMP = tempfile.mkdtemp(prefix="leady_v3_test_")
os.environ["DATA_DIR"] = TMP

import app as A                      # noqa: E402
import calendar_view as cv           # noqa: E402
import db                            # noqa: E402
import repo                          # noqa: E402
from seed import bootstrap           # noqa: E402

KL = A.app.test_client()

# --- logowanie w testach ---------------------------------------------------
# Od v5 aplikacja wymaga konta i tokenu CSRF. Testy sprawdzają logikę biznesową,
# nie ekran logowania (ten ma własny plik), więc zakładamy konto koordynatora
# i logujemy klienta raz, na starcie.
def _zaloguj_testowo():
    import db as _db, uzytkownicy as _uz
    c = _db.get_conn()
    _uz.init(c)
    if not _uz.znajdz(c, "TEST-koordynator"):
        _uz.utworz(c, "TEST-koordynator", "koordynator", "1379")
    c.close()
    r = KL.post("/api/logowanie", json={"osoba": "TEST-koordynator", "pin": "1379"})
    assert r.status_code == 200, "logowanie testowe nie przeszło: %s" % r.get_data()
    with KL.session_transaction() as s:
        s["csrf"] = "test-csrf"
    KL.environ_base["HTTP_X_CSRF"] = "test-csrf"

WYNIKI = []


def sprawdz(nazwa, warunek, opis=""):
    WYNIKI.append((nazwa, bool(warunek), opis))
    znak = "OK  " if warunek else "BLAD"
    print("  [%s] %s%s" % (znak, nazwa, (" — " + opis) if opis else ""))
    return bool(warunek)


def patch_lead(lead_id, field, value):
    r = KL.patch("/api/lead/%d" % lead_id,
                 data=json.dumps({"field": field, "value": value}),
                 content_type="application/json")
    return r.status_code, r.get_json()


def post(url, payload):
    r = KL.post(url, data=json.dumps(payload), content_type="application/json")
    return r.status_code, r.get_json()


def leady(**filtry):
    conn = db.get_conn()
    f = repo.pusty_filtr()
    f.update(filtry)
    rows = repo.filtruj_leady(conn, f)
    conn.close()
    return rows


def nowa_szkola(nazwa, miasto="08. Katowice", typ="01. Szkoła podstawowa"):
    kod, dane = post("/api/lead", {"nazwa": nazwa, "miejscowosc": miasto, "typ": typ})
    assert kod == 200 and dane.get("ok"), (kod, dane)
    return dane["id"]


# =====================================================================

def main():
    print("Baza testowa:", TMP)
    bootstrap()
    _zaloguj_testowo()

    print("\nS0 — start: pusta baza, gotowe słowniki")
    conn = db.get_conn()
    sprawdz("słownik handlowców niepusty", len(db.slownik_values(conn, "handlowiec")) >= 5)
    sprawdz("słownik trenerów niepusty", len(db.slownik_values(conn, "trener")) >= 30)
    m0 = repo.metryki(conn)
    conn.close()
    sprawdz("zero leadów na starcie", m0["leady"] == 0)

    # -----------------------------------------------------------------
    print("\nS1 — Koordynator przypisuje szkołę handlowcowi i daje termin")
    lead_id = nowa_szkola("SP 11 Będzin", miasto="15. Będzin")
    sprawdz("szkoła jest na liście do rozdania",
            any(r["id"] == lead_id for r in leady(zakres="nieprzydzielone")))
    kod, _ = post("/api/przypisz", {"ids": [lead_id], "handlowiec": "04. Chytry",
                                    "deadline": "2026-09-30"})
    sprawdz("przypisanie się udało", kod == 200)
    sprawdz("zniknęła z listy do rozdania",
            not any(r["id"] == lead_id for r in leady(zakres="nieprzydzielone")),
            "to filtr, nie usunięcie wiersza")
    u_chytrego = leady(handlowiec="04. Chytry")
    sprawdz("jest u Chytrego", any(r["id"] == lead_id for r in u_chytrego))
    sprawdz("ma wpisany termin ostateczny",
            u_chytrego[0]["deadline"] == "2026-09-30")

    # -----------------------------------------------------------------
    print("\nS1b — „Przedłuż termin”: +N dni od terminu, po terminie od dziś")
    kod, j = post("/api/przedluz", {"ids": [lead_id], "dni": 14})
    sprawdz("przedłużenie przechodzi", kod == 200 and (j or {}).get("n") == 1)
    row = [r for r in leady() if r["id"] == lead_id][0]
    sprawdz("termin przesunięty o 14 dni od poprzedniego",
            row["deadline"] == "2026-10-14", str(row["deadline"]))

    # Lead dawno po terminie: liczymy od DZIŚ — licząc od starej daty przycisk
    # dawałby datę nadal w przeszłości i automat zabrałby szkołę mimo przedłużenia.
    conn = db.get_conn()
    conn.execute("UPDATE leady SET deadline='2026-01-01' WHERE id=?", (lead_id,))
    conn.commit(); conn.close()
    post("/api/przedluz", {"ids": [lead_id], "dni": 7})
    oczekiwany = (dt.date.today() + dt.timedelta(days=7)).isoformat()
    row = [r for r in leady() if r["id"] == lead_id][0]
    sprawdz("po terminie liczy od dziś — data nie ląduje w przeszłości",
            row["deadline"] == oczekiwany, str(row["deadline"]))
    conn = db.get_conn()
    ost = conn.execute("SELECT co FROM log WHERE lead_id=? ORDER BY id DESC LIMIT 1",
                       (lead_id,)).fetchone()
    # dalsze scenariusze liczą na termin z S1 — przywracamy go
    conn.execute("UPDATE leady SET deadline='2026-09-30' WHERE id=?", (lead_id,))
    conn.commit(); conn.close()
    sprawdz("przedłużenie zostawia ślad w historii",
            ost is not None and ost["co"] == "przedłużenie terminu")

    # -----------------------------------------------------------------
    print("\nS2 — Handlowiec umawia DT: trafia do Zbiorczego i do kalendarza")
    kod, dane = post("/api/event", {
        "lead_id": lead_id, "typ": "DT", "data": "2026-09-16",
        "godz_od": "08:00", "godz_do": "09:35", "trener": "01. Małolepsza",
        "ilosc_klas": 10, "ilosc_dzieci": 186, "numer_sali": "12"})
    sprawdz("DT dodane", kod == 200 and dane.get("ok"))
    sprawdz("brak ostrzeżenia o kolizji przy pierwszym DT", dane.get("kolizja") is None)
    row = [r for r in leady() if r["id"] == lead_id][0]
    sprawdz("status sam przeszedł na „DT umówione”",
            row["status_realizacji"] == "03. DT umówione", row["status_realizacji"])
    sprawdz("jest w Zbiorczym (widok „umówione”)",
            any(r["id"] == lead_id for r in leady(zakres="umowione")))
    conn = db.get_conn()
    mx = cv.build_matrix(conn, "2026-09")
    conn.close()
    komorki = [c for t in mx["tygodnie"] for w in t["wiersze"]
               if w["trener"] == "01. Małolepsza" for c in w["cells"] if c]
    sprawdz("widać wpis w kalendarzu wrześniowym", len(komorki) == 1)

    # skok do daty (08.08): data w adresie wygrywa z miesiącem i podświetla tydzień
    r = KL.get("/kalendarz?d=2026-09-16&m=2026-08")
    html_kal = r.get_data(as_text=True)
    sprawdz("skok do daty otwiera właściwy miesiąc mimo innego `m`",
            r.status_code == 200 and "2026-09" in html_kal)
    sprawdz("tydzień z wybraną datą jest podświetlony", "tydzien-wybrany" in html_kal)
    sprawdz("zła data nie wysypuje kalendarza",
            KL.get("/kalendarz?d=krzak").status_code == 200)

    # REGRESJA (zgłoszone 09.08): wpisanie „0002" zamiast „2026" przenosiło
    # kalendarz do roku 2 n.e., a lista miesięcy ma tylko miesiące z danymi —
    # nie było czym wrócić. Data spoza widełek ma być IGNOROWANA, nie honorowana.
    r = KL.get("/kalendarz?d=0002-08-09")
    h = r.get_data(as_text=True)
    sprawdz("rok 2 n.e. w polu daty nie zabiera kalendarza w przeszłość",
            r.status_code == 200 and "0002-" not in h)
    sprawdz("po złej dacie kalendarz stoi na sensownym miesiącu",
            "2026-" in h)
    r = KL.get("/kalendarz?m=0002-08")
    sprawdz("to samo dla miesiąca wpisanego wprost w adres",
            r.status_code == 200 and "0002-" not in r.get_data(as_text=True))
    sprawdz("dostępność też ma bezpiecznik na miesiąc",
            "0002-" not in KL.get("/dostepnosc?m=0002-08").get_data(as_text=True))
    sprawdz("pole daty ogranicza rok w przeglądarce",
            'min="2025-01-01"' in html_kal and 'max="2035-12-31"' in html_kal)
    sprawdz("rok 3000 też odrzucony",
            "3000-" not in KL.get("/kalendarz?d=3000-01-01").get_data(as_text=True))

    # Zgłoszone 09.08: wybór miesiąca ginął przy przejściu kalendarz ↔ dostępność
    # (linki w nawigacji nie niosą `m`, więc każdy ekran startował od ostatniego
    # miesiąca z danymi). Wybór ma przeżyć zmianę ekranu.
    KL.get("/kalendarz?m=2026-09")
    sprawdz("dostępność otwiera się na miesiącu wybranym w kalendarzu",
            "2026-09" in KL.get("/dostepnosc").get_data(as_text=True))
    KL.get("/dostepnosc?m=2026-10")
    sprawdz("i odwrotnie — kalendarz pamięta wybór z dostępności",
            "2026-10" in KL.get("/kalendarz").get_data(as_text=True))
    sprawdz("adres z konkretnym miesiącem dalej wygrywa",
            "2026-09" in KL.get("/kalendarz?m=2026-09").get_data(as_text=True))

    # -----------------------------------------------------------------
    print("\nS3 — DWA DT jednego trenera w jednym dniu (ZGLOSZONY BUG)")
    lead2 = nowa_szkola("SP 8 Będzin", miasto="15. Będzin")
    post("/api/przypisz", {"ids": [lead2], "handlowiec": "02. Olszewska",
                           "deadline": "2026-09-30"})
    kod, dane2 = post("/api/event", {
        "lead_id": lead2, "typ": "DT", "data": "2026-09-16",
        "godz_od": "11:00", "godz_do": "12:30", "trener": "01. Małolepsza"})
    sprawdz("drugie DT tego samego dnia przyjęte", kod == 200 and dane2.get("ok"),
            "zapisu NIE blokujemy")
    sprawdz("brak ostrzeżenia — godziny się nie nakładają", dane2.get("kolizja") is None)
    conn = db.get_conn()
    mx = cv.build_matrix(conn, "2026-09")
    conn.close()
    komorka = None
    for t in mx["tygodnie"]:
        for w in t["wiersze"]:
            if w["trener"] != "01. Małolepsza":
                continue
            for i, c in enumerate(w["cells"]):
                if t["dni"][i]["iso"] == "2026-09-16":
                    komorka = c
    sprawdz("W JEDNEJ KOMÓRCE KALENDARZA SĄ DWA WPISY",
            komorka is not None and len(komorka) == 2,
            "w arkuszu XLOOKUP pokazywał tylko pierwszy")
    if komorka:
        godziny = sorted(e["godz_od"] for e in komorka)
        sprawdz("oba wpisy z właściwymi godzinami", godziny == ["08:00", "11:00"],
                " / ".join(godziny))

    # -----------------------------------------------------------------
    print("\nS4 — Trzecie DT NAKŁADA się godzinami: ostrzeżenie, ale zapis przechodzi")
    lead3 = nowa_szkola("SP 20 Sosnowiec", miasto="13. Sosnowiec")
    post("/api/przypisz", {"ids": [lead3], "handlowiec": "04. Chytry",
                           "deadline": "2026-09-15"})
    kod, dane3 = post("/api/event", {
        "lead_id": lead3, "typ": "DT", "data": "2026-09-16",
        "godz_od": "09:00", "godz_do": "10:00", "trener": "01. Małolepsza"})
    sprawdz("zapis przeszedł", kod == 200 and dane3.get("ok"))
    sprawdz("aplikacja OSTRZEGŁA o nakładaniu", bool(dane3.get("kolizja")),
            dane3.get("kolizja") or "")
    conn = db.get_conn()
    kol = cv.lista_kolizji(conn, "2026-09")
    conn.close()
    sprawdz("kolizja widoczna na pulpicie", len(kol) == 2,
            "dwa nakładające się wpisy: 08:00–09:35 i 09:00–10:00")

    # -----------------------------------------------------------------
    print("\nS5 — Przesunięcie terminu: poprawiam w jednym miejscu")
    conn = db.get_conn()
    ev = conn.execute("SELECT id FROM eventy WHERE lead_id=? AND typ='DT'",
                      (lead_id,)).fetchone()["id"]
    conn.close()
    r = KL.patch("/api/event/%d" % ev,
                 data=json.dumps({"field": "data", "value": "2026-09-23"}),
                 content_type="application/json")
    sprawdz("data zmieniona", r.status_code == 200)
    conn = db.get_conn()
    mx = cv.build_matrix(conn, "2026-09")
    conn.close()
    znalezione = {}
    for t in mx["tygodnie"]:
        for w in t["wiersze"]:
            if w["trener"] != "01. Małolepsza":
                continue
            for i, c in enumerate(w["cells"]):
                if c:
                    znalezione[t["dni"][i]["iso"]] = len(c)
    sprawdz("wpis przeskoczył na 23.09", znalezione.get("2026-09-23") == 1, str(znalezione))
    sprawdz("pod 16.09 zostały dwa", znalezione.get("2026-09-16") == 2)
    sprawdz("kolizja zniknęła sama", len(cv.lista_kolizji(db.get_conn(), "2026-09")) == 0)

    # -----------------------------------------------------------------
    print("\nS6 — Nowy miesiąc tworzy się sam (bez konfigurowania zakładki)")
    conn = db.get_conn()
    przed = cv.available_months(conn)
    conn.close()
    sprawdz("listopada jeszcze nie ma", "2026-11" not in przed, str(przed))
    post("/api/event", {"lead_id": lead3, "typ": "DT", "data": "2026-11-03",
                        "godz_od": "08:00", "godz_do": "09:35",
                        "trener": "04. Zemela"})
    conn = db.get_conn()
    po = cv.available_months(conn)
    mx11 = cv.build_matrix(conn, "2026-11")
    conn.close()
    sprawdz("listopad pojawił się sam", "2026-11" in po)
    sprawdz("i ma w sobie wpis", mx11["n_events"] == 1)

    # -----------------------------------------------------------------
    print("\nS7 — Odebranie leada: wraca do puli, nic nie ginie")
    kod, _ = post("/api/odbierz", {"ids": [lead3]})
    sprawdz("odebranie się udało", kod == 200)
    niew = leady(zakres="niewykorzystane")
    sprawdz("lead jest w „Niewykorzystane rekordy”",
            any(r["id"] == lead3 for r in niew))
    row3 = [r for r in leady() if r["id"] == lead3][0]
    sprawdz("nie ma już handlowca", not row3["handlowiec"])
    sprawdz("dane placówki nietknięte", row3["placowka"] == "SP 20 Sosnowiec")
    kod, _ = post("/api/przypisz", {"ids": [lead3], "handlowiec": "02. Olszewska",
                                    "deadline": "2026-10-15"})
    row3 = [r for r in leady() if r["id"] == lead3][0]
    sprawdz("przydzielony innemu handlowcowi", row3["handlowiec"] == "02. Olszewska")
    sprawdz("wypadł z puli zwrotnej",
            not any(r["id"] == lead3 for r in leady(zakres="niewykorzystane")))

    # -----------------------------------------------------------------
    print("\nS8 — Lista rozwijana wymuszona: koniec z „02. Olaszewska”")
    kod, dane = patch_lead(lead_id, "handlowiec", "02. Olaszewska")
    sprawdz("literówka ODRZUCONA", kod == 400 and not dane.get("ok"),
            dane.get("error", ""))
    kod, _ = patch_lead(lead_id, "handlowiec", "02. Olszewska")
    sprawdz("poprawna wartość przyjęta", kod == 200)
    kod, dane = patch_lead(lead_id, "status_realizacji", "cokolwiek")
    sprawdz("status spoza listy odrzucony", kod == 400)

    # -----------------------------------------------------------------
    print("\nS9 — „Wybrane szkoły na tydzień do góry”")
    kod, dane = post("/api/pin", {"id": lead2, "pin": True})
    sprawdz("przypięcie zapisane", kod == 200 and dane.get("pin"))
    przypiete = leady(zakres="pin")
    sprawdz("jest na planie tygodnia", any(r["id"] == lead2 for r in przypiete))
    wszystkie = leady()
    sprawdz("przypięty jest NA GÓRZE listy", wszystkie[0]["id"] == lead2,
            "sortowanie stawia przypięte pierwsze")
    r = KL.get("/tydzien")
    sprawdz("ekran „Tydzień” się otwiera", r.status_code == 200)

    # -----------------------------------------------------------------
    print("\nS10 — Zajęcia cykliczne: jedna reguła, wiele wystąpień")
    kod, dane = post("/api/event", {
        "lead_id": lead2, "typ": "CYKLICZNE", "data": "2026-10-05",
        "godz_od": "12:30", "godz_do": "13:30", "trener": "14. Swoboda",
        "cykl_dzien": "poniedziałek", "co_ile_tygodni": 1, "grupa": "1",
        "sprzet": "01. Sala komputerowa", "kod_tinkercad": "BMR DKP QHW"})
    sprawdz("cykl dodany jako JEDEN rekord", kod == 200 and dane.get("ok"))
    conn = db.get_conn()
    ile_rekordow = conn.execute("SELECT COUNT(*) c FROM eventy "
                                "WHERE typ='CYKLICZNE'").fetchone()["c"]
    ag = cv.build_agenda(conn, "2026-10", typy=["CYKLICZNE"])
    conn.close()
    sprawdz("w bazie jest 1 rekord", ile_rekordow == 1)
    sprawdz("a w październiku widać 4 wystąpienia", ag["n_events"] == 4,
            "%d wystąpień" % ag["n_events"])

    print("\nS11 — Zastępstwo na jednej dacie = wyjątek, nie kopiowanie tygodnia")
    conn = db.get_conn()
    eid = conn.execute("SELECT id FROM eventy WHERE typ='CYKLICZNE'").fetchone()["id"]
    conn.execute("INSERT INTO wyjatki_cyklu (event_id, data, zastepstwo, uwagi) "
                 "VALUES (?,?,?,?)", (eid, "2026-10-19", "09. Bochniarz", "zastępstwo"))
    conn.commit()
    ag = cv.build_agenda(conn, "2026-10", typy=["CYKLICZNE"])
    conn.close()
    dzien19 = [d for d in ag["dni"] if d["iso"] == "2026-10-19"]
    sprawdz("19.10 nadal ma zajęcia", bool(dzien19))
    if dzien19:
        e = dzien19[0]["eventy"][0]
        sprawdz("na tej dacie jest zastępstwo", e.get("zastepstwo") == "09. Bochniarz")
        sprawdz("prowadzący reguły bez zmian", e.get("trener") == "14. Swoboda")

    print("\nS12 — Odwołane zajęcia znikają tylko z tej jednej daty")
    conn = db.get_conn()
    conn.execute("INSERT INTO wyjatki_cyklu (event_id, data, odwolane, uwagi) "
                 "VALUES (?,?,1,?)", (eid, "2026-10-26", "odwołane"))
    conn.commit()
    ag = cv.build_agenda(conn, "2026-10", typy=["CYKLICZNE"])
    conn.close()
    sprawdz("zostały 3 wystąpienia z 4", ag["n_events"] == 3)
    sprawdz("26.10 nie ma zajęć",
            not any(d["iso"] == "2026-10-26" for d in ag["dni"]))

    # -----------------------------------------------------------------
    print("\nS13 — Eksport oddaje DOKŁADNIE to, co widać po filtrach")
    from exporter import build_workbook
    import openpyxl
    conn = db.get_conn()
    f = repo.pusty_filtr(); f["zakres"] = "umowione"
    rows_f = repo.filtruj_leady(conn, f)
    bio = build_workbook(conn, rows_f, f)
    conn.close()
    wb = openpyxl.load_workbook(bio)
    ws = wb["Leady"]
    sprawdz("arkusze zgodne z zapowiedzią",
            set(["Leady", "Spotkania", "Kolizje", "Filtr"]).issubset(set(wb.sheetnames)),
            ", ".join(wb.sheetnames))
    sprawdz("liczba wierszy = liczba wyfiltrowanych",
            ws.max_row - 1 == len(rows_f), "%d w pliku / %d na ekranie"
            % (ws.max_row - 1, len(rows_f)))
    sprawdz("nagłówki w języku klienta", ws.cell(1, 1).value == "Handlowiec")
    r = KL.get("/export.xlsx?zakres=umowione")
    sprawdz("pobranie przez przeglądarkę działa", r.status_code == 200)

    # -----------------------------------------------------------------
    print("\nS14 — Historia zmian przy leadzie (kontrola aktywności)")
    conn = db.get_conn()
    lead = repo.lead_szczegoly(conn, lead_id)
    conn.close()
    sprawdz("log zapisał zmiany", len(lead["log"]) >= 3, "%d wpisów" % len(lead["log"]))
    sprawdz("jest ślad przypisania",
            any(w["co"] == "przypisanie" for w in lead["log"]))
    sprawdz("zapisana ostatnia aktywność", bool(lead["ostatnia_aktywnosc"]))

    print("\nS15 — Wszystkie ekrany otwierają się na tych danych")
    for s in ["/pulpit", "/baza", "/leady", "/zbiorczy", "/niewykorzystane",
              "/tydzien", "/slowniki", "/import", "/kalendarz",
              "/kalendarz?widok=agenda", "/kalendarz?widok=starty",
              "/lead/%d" % lead_id]:
        sprawdz("otwiera się %s" % s, KL.get(s).status_code == 200)

    # -----------------------------------------------------------------
    print("\nS16 — Zajęcia cykliczne widać w kalendarzu (regresja 10.08)")
    #
    # Na produkcji dwa jedyne wpisy cykliczne były NIEWIDOCZNE w każdym miesiącu.
    # Przyczyna: openpyxl oddaje komórkę z datą jako `datetime`, więc wyliczona
    # data pierwszych zajęć trafiała do bazy jako „2026-09-22T00:00:00".
    # `date.fromisoformat` odrzuca taki zapis, a kalendarz robił wtedy `continue`
    # — czyli pomijał wpis BEZ ŚLADU. Test pilnuje obu stron naprawy: importer
    # ma zapisywać samą datę, a kalendarz ma sobie radzić z tym, co już leży
    # w danych (na produkcji nie zrobimy importu od nowa).
    import calendar_view as cal
    import importer as imp

    d = imp._pierwsza_data_dnia("wtorek", dt.datetime(2026, 9, 15, 0, 0))
    sprawdz("wyliczona data to data, nie znacznik czasu",
            isinstance(d, dt.date) and not isinstance(d, dt.datetime), repr(d))
    sprawdz("i wypada we właściwym dniu tygodnia", d.weekday() == 1, str(d))

    conn = db.get_conn()
    lid = conn.execute("SELECT id FROM leady LIMIT 1").fetchone()[0]
    # celowo w starym, zepsutym formacie — tak wygląda to na produkcji
    conn.execute("INSERT INTO eventy (lead_id, typ, data, cykl_dzien, godz_od) "
                 "VALUES (?, 'CYKLICZNE', '2026-09-22T00:00:00', 'wtorek', '13:30')", (lid,))
    conn.commit()
    wrzesien = cal.events_for_month(conn, "2026-09", typy=("CYKLICZNE",))
    pazdziernik = cal.events_for_month(conn, "2026-10", typy=("CYKLICZNE",))
    conn.close()
    sprawdz("kalendarz nie pomija wpisu ze znacznikiem czasu",
            len(wrzesien) >= 1, "%d wystąpień we wrześniu" % len(wrzesien))
    sprawdz("cykl rozwija się na kolejny miesiąc",
            len(pazdziernik) >= 3, "%d wystąpień w październiku" % len(pazdziernik))
    sprawdz("każde wystąpienie ma czystą datę",
            all(len(e["data"]) == 10 for e in wrzesien + pazdziernik))

    print("\nS17 — Zajęcia bez prowadzącego są WIDOCZNE w macierzy (regresja 10.08)")
    #
    # Wiersze macierzy to trenerzy, więc zajęcia bez przypisanej osoby nie miały
    # gdzie się pokazać i znikały z widoku domyślnego — a licznik u góry i tak
    # je liczył. Na danych klienta: 56 pokazanych z 61 zapowiedzianych.
    # To ten sam błąd, przed którym kod broni się przy kolizjach: schowanie
    # czegoś po cichu jest gorsze niż pokazanie brzydko.
    conn = db.get_conn()
    lid2 = conn.execute("SELECT id FROM leady LIMIT 1").fetchone()[0]
    conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od) "
                 "VALUES (?, 'DT', '2026-09-17', '09:00')", (lid2,))   # bez trenera
    conn.commit()
    mx = cal.build_matrix(conn, "2026-09", tylko_zajete=True)
    ile_w_macierzy = sum(len(cc) for t in mx["tygodnie"]
                         for w in t["wiersze"] for cc in w["cells"])
    conn.close()
    sprawdz("macierz pokazuje tyle, ile zapowiada licznik",
            ile_w_macierzy == mx["n_events"],
            "%d pokazanych / %d w liczniku" % (ile_w_macierzy, mx["n_events"]))
    sprawdz("jest wiersz zbiorczy dla braku prowadzącego",
            cal.BEZ_TRENERA in mx["trenerzy"])
    sprawdz("i stoi na samej górze, nie między trenerami",
            mx["trenerzy"][0] == cal.BEZ_TRENERA)
    sprawdz("licznik braków trafia do widoku", mx["n_bez_trenera"] >= 1,
            "%d bez prowadzącego, %d bez godziny"
            % (mx["n_bez_trenera"], mx["n_bez_godziny"]))
    sprawdz("wiersz jest oznaczony flagą, nie samą nazwą",
            any(w.get("brak_trenera") for t in mx["tygodnie"] for w in t["wiersze"]))

    # =================================================================
    # S18 — „chcę zobaczyć DT razem z przedszkolami, ale bez cykli szkolnych"
    #
    # Filtr typu miał trzy pozycje i jedna z nich kłamała: „— DT i cykliczne —"
    # pokazywało WSZYSTKO, łącznie z festynami i wpisami VR. Po dołożeniu
    # wariantu przedszkolnego doszła potrzeba par (DT + jeden rodzaj cyklu),
    # bo koordynatorka planuje osobno szkoły i osobno przedszkola.
    print("\nS18 — filtr typu w kalendarzu: sześć pozycji, każda dosłowna")
    conn = db.get_conn()
    pid = conn.execute("INSERT INTO placowki (nazwa, miejscowosc) VALUES (?,?)",
                       ("Placówka S18", "Gliwice")).lastrowid
    lid = conn.execute("INSERT INTO leady (placowka_id) VALUES (?)", (pid,)).lastrowid
    for typ, data in (("DT", "2027-03-02"), ("CYKLICZNE", "2027-03-03"),
                      ("CYKLICZNE-PRZEDSZKOLE", "2027-03-04"), ("FESTYN", "2027-03-05")):
        eid = conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od, trener) "
                           "VALUES (?,?,?,?,?)",
                           (lid, typ, data, "09:00", "01. Nowak")).lastrowid
        if typ == "CYKLICZNE-PRZEDSZKOLE":
            # pakiet jednodniowy — żeby cykl nie rozwinął się na kolejne miesiące
            conn.execute("INSERT INTO terminy_cyklu (event_id, nr, data) VALUES (?,1,?)",
                         (eid, data))
    conn.commit()

    def typy_widoczne(klucz):
        _, typy = A._typy_kalendarza({"typ": klucz})
        return sorted({e["typ"] for e in cv.events_for_month(conn, "2027-03", typy=typy)})

    sprawdz("„wszystko” pokazuje też festyn (nazwa nie kłamie)",
            typy_widoczne("") == ["CYKLICZNE", "CYKLICZNE-PRZEDSZKOLE", "DT", "FESTYN"])
    sprawdz("tylko DT", typy_widoczne("DT") == ["DT"])
    # Do 17.08 „CYKLICZNE" znaczyło OBA warianty cyklu. Rozdzielone na prośbę
    # klienta: szkoły i przedszkola planuje się osobno.
    sprawdz("tylko CYKLICZNE — bez przedszkoli",
            typy_widoczne("CYKLICZNE") == ["CYKLICZNE"])
    sprawdz("tylko CYKLICZNE-PRZEDSZKOLE",
            typy_widoczne("CYKLICZNE-PRZEDSZKOLE") == ["CYKLICZNE-PRZEDSZKOLE"])
    sprawdz("DT i CYKLICZNE", typy_widoczne("DT,CYKLICZNE") == ["CYKLICZNE", "DT"])
    sprawdz("DT i CYKLICZNE-PRZEDSZKOLE",
            typy_widoczne("DT,CYKLICZNE-PRZEDSZKOLE") == ["CYKLICZNE-PRZEDSZKOLE", "DT"])
    conn.close()

    # Wartość spoza listy ma otworzyć kalendarz, a nie pusty ekran — stara
    # zakładka z czasów, gdy filtr miał inne wartości, dalej ma działać.
    sprawdz("nieznana wartość filtra = wszystko",
            A._typy_kalendarza({"typ": "COKOLWIEK"}) == ("", None))
    # W adresie `+` to zakodowana spacja. Gdyby pary rozdzielał plus, link
    # wklejony z notatki przychodziłby jako „DT CYKLICZNE" i cicho przestawał
    # filtrować — dlatego rozdziela przecinek.
    sprawdz("pary rozdziela przecinek, nie plus",
            all("+" not in k for k in A.FILTRY_TYPU_MAPA))

    r = KL.get("/kalendarz?m=2027-03")
    html = r.get_data(as_text=True)
    sprawdz("kalendarz z nowym filtrem otwiera się", r.status_code == 200)
    sprawdz("wszystkie sześć pozycji jest na liście",
            all(etykieta in html for _, etykieta, _ in A.FILTRY_TYPU))
    sprawdz("wybrana pozycja zostaje zaznaczona po przeładowaniu",
            'value="DT,CYKLICZNE" selected'
            in KL.get("/kalendarz?m=2027-03&typ=DT,CYKLICZNE").get_data(as_text=True))

    # --- P05: na czym otwiera się kalendarz (zgłoszenie K11) ---------------
    #
    # „kalendarz ustawia się na czerwiec na starcie a nie na wrzesień" — Kasia,
    # 20.08. Dwie przyczyny, obie warte testu: fallback brał miesiąc NAJDALSZY
    # w przyszłość (przy cyklach sięgających pół roku to miesiąc, w którym nikt
    # nie pracuje), a zapamiętany w sesji wybór nie miał daty ważności.
    print("\n-- P05: domyślny miesiąc kalendarza --")
    prawdziwe_dzis = A.dzis
    try:
        A.dzis = lambda: "2026-08-20"
        sprawdz("pusty sierpień → skacze na wrzesień, nie na październik",
                A._miesiac_domyslny(["2026-07", "2026-09", "2026-10"]) == "2026-09")
        sprawdz("bieżący miesiąc wygrywa, gdy coś w nim jest",
                A._miesiac_domyslny(["2026-07", "2026-08", "2026-09"]) == "2026-08")
        sprawdz("same przeszłe miesiące → ostatni z nich",
                A._miesiac_domyslny(["2026-05", "2026-06"]) == "2026-06")
        sprawdz("pusta baza → bieżący miesiąc",
                A._miesiac_domyslny([]) == "2026-08")

        # Sedno zgłoszenia: jedno zajrzenie do czerwca nie może zostać na 30 dni.
        with KL.session_transaction() as s:
            s["miesiac"] = "2026-06"
        r = KL.get("/kalendarz")
        sprawdz("kalendarz się otwiera", r.status_code == 200, "kod %s" % r.status_code)
        with KL.session_transaction() as s:
            sprawdz("zapamiętany czerwiec został skasowany z sesji",
                    s.get("miesiac") != "2026-06", s.get("miesiac"))

        # ...ale miesiąc PRZYSZŁY dalej przeżywa przejście na sąsiedni ekran
        # (to była poprawka z 09.08 i nie wolno jej odkręcić).
        with KL.session_transaction() as s:
            s["miesiac"] = "2026-10"
        KL.get("/kalendarz")
        with KL.session_transaction() as s:
            sprawdz("przyszły miesiąc dalej jest pamiętany",
                    s.get("miesiac") == "2026-10", s.get("miesiac"))
    finally:
        A.dzis = prawdziwe_dzis
        with KL.session_transaction() as s:
            s.pop("miesiac", None)

    # --- P24: filtr „bez prowadzącego" (pytanie Zuzi 20.08) -----------------
    #
    # „czy można już wyszukiwać bez prowadzącego?" — do 20.08 nie. Licznik
    # w nagłówku pokazywał, ile takich zajęć jest, ale nie dało się do nich
    # dojść: filtr na chipach szuka WPISANEGO TEKSTU, a brak prowadzącego to
    # brak wartości, którego żadnym fragmentem nie da się wpisać.
    print("\n-- P24: filtr „bez prowadzącego” --")
    sprawdz("sam filtr: zostają tylko wpisy bez prowadzącego",
            [e["id"] for e in cv.tylko_bez_obsady(
                [{"id": 1, "trener": "13. Cebula"}, {"id": 2, "trener": ""},
                 {"id": 3, "trener": None}, {"id": 4, "trener": "   "}])] == [2, 3, 4])
    sprawdz("drukarz nie zastępuje prowadzącego",
            len(cv.tylko_bez_obsady([{"trener": "", "drukarz": "04. Zemela"}])) == 1)

    conn = db.get_conn()
    cur = conn.execute("INSERT INTO placowki (nazwa, zrodlo) VALUES "
                       "('SP WOLNA do obsadzenia', 'test')")
    l_wolna = conn.execute("INSERT INTO leady (placowka_id) VALUES (?)",
                           (cur.lastrowid,)).lastrowid
    cur = conn.execute("INSERT INTO placowki (nazwa, zrodlo) VALUES "
                       "('SP OBSADZONA przez trenera', 'test')")
    l_obsadzona = conn.execute("INSERT INTO leady (placowka_id) VALUES (?)",
                               (cur.lastrowid,)).lastrowid
    conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od) "
                 "VALUES (?, 'DT', '2026-11-04', '09:00')", (l_wolna,))
    conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od, trener) "
                 "VALUES (?, 'DT', '2026-11-05', '09:00', '13. Cebula')", (l_obsadzona,))
    conn.commit()
    conn.close()

    for widok in ("macierz", "agenda", "starty"):
        r = KL.get("/kalendarz?m=2026-11&widok=%s&bez_obsady=1" % widok)
        html = r.get_data(as_text=True)
        sprawdz("%s: zostaje szkoła bez prowadzącego" % widok,
                "SP WOLNA do obsadzenia" in html, "kod %s" % r.status_code)
        sprawdz("%s: znika szkoła z prowadzącym" % widok,
                "SP OBSADZONA przez trenera" not in html)

    r = KL.get("/kalendarz?m=2026-11&widok=agenda")
    html = r.get_data(as_text=True)
    sprawdz("bez filtra widać obie", "SP WOLNA do obsadzenia" in html
            and "SP OBSADZONA przez trenera" in html)
    sprawdz("licznik braków prowadzi do filtra", "bez_obsady=1" in html)

    # --- P08: odwołanie DT ze śladem (zgłoszenie K12) -----------------------
    #
    # „nie widzę też możliwości wykasowania czegoś z kalendarza, w razie jakby
    # np. szkoła w ostatnim momencie odmówiła współpracy" — Kasia, 20.08.
    #
    # Odwołujemy ZE ŚLADEM zamiast kasować: wpis zostaje w bazie jako dowód, że
    # temat był i się nie udał (raport wykonania liczy właśnie takie przypadki),
    # ale znika z grafiku i przestaje zajmować trenerowi termin.
    print("\n-- P08: odwołanie DT --")
    l_odw = nowa_szkola("SP 77 do odwołania", miasto="15. Będzin")
    post("/api/przypisz", {"ids": [l_odw], "handlowiec": "04. Chytry"})
    kod, _ = post("/api/event", {"lead_id": l_odw, "typ": "DT", "data": "2026-12-08",
                                 "godz_od": "09:00", "trener": "01. Małolepsza"})
    sprawdz("DT do odwołania dodane", kod == 200)
    row = [r for r in leady() if r["id"] == l_odw][0]
    sprawdz("status wskoczył na sukces", row["status_realizacji"].startswith("03."))

    conn = db.get_conn()
    ev_id = conn.execute("SELECT id FROM eventy WHERE lead_id=?", (l_odw,)).fetchone()["id"]
    conn.close()

    kod, j = post("/api/event/%d/odwolaj" % ev_id, {"powod": "szkoła wycofała się dzień przed"})
    sprawdz("odwołanie przechodzi", kod == 200, str(j)[:90])
    sprawdz("aplikacja mówi, że szkoła wróciła do umawiania",
            (j or {}).get("wrocil_do_umawiania") is True)

    conn = db.get_conn()
    w_grafiku = [e for e in cv.events_for_month(conn, "2026-12") if e["lead_id"] == l_odw]
    conn.close()
    sprawdz("odwołane zajęcia znikają z grafiku", not w_grafiku)

    row = [r for r in leady() if r["id"] == l_odw][0]
    sprawdz("lead przestał być domknięty",
            not (row["status_realizacji"] or "").startswith("03."), row["status_realizacji"])
    sprawdz("termin DT zniknął z karty leada", not row["dt_data"], str(row["dt_data"]))
    sprawdz("szkoła NIE wróciła do puli — handlowiec zostaje",
            row["handlowiec"] == "04. Chytry", str(row["handlowiec"]))

    conn = db.get_conn()
    ile = conn.execute("SELECT COUNT(*) c FROM eventy WHERE id=?", (ev_id,)).fetchone()["c"]
    powod = conn.execute("SELECT powod_odwolania p FROM eventy WHERE id=?",
                         (ev_id,)).fetchone()["p"]
    conn.close()
    sprawdz("wpis został w bazie jako dowód", ile == 1)
    sprawdz("powód zapisany przy wpisie", "wycofała" in (powod or ""))

    # Pulpit i kalendarz muszą mówić to samo. Licznik, który liczy odwołane,
    # a grafik, który ich nie pokazuje, to dwie liczby i żadnej wiadomo, która
    # kłamie — a to najgorszy rodzaj błędu w tym projekcie.
    conn = db.get_conn()
    m = repo.metryki(conn)
    wszystkich_dt = conn.execute(
        "SELECT COUNT(*) c FROM eventy WHERE typ='DT'").fetchone()["c"]
    czynnych_dt = conn.execute(
        "SELECT COUNT(*) c FROM eventy WHERE typ='DT' "
        "AND (odwolane IS NULL OR odwolane='')").fetchone()["c"]
    conn.close()
    sprawdz("pulpit liczy tylko czynne DT, nie wszystkie wiersze",
            m["eventy_dt"] == czynnych_dt and czynnych_dt < wszystkich_dt,
            "pulpit %d, czynnych %d, w bazie %d" % (m["eventy_dt"], czynnych_dt,
                                                    wszystkich_dt))
    sprawdz("pulpit pokazuje odwołane osobno", m["eventy_odwolane"] >= 1,
            "odwołanych: %s" % m["eventy_odwolane"])

    kod, _ = post("/api/event/%d/odwolaj" % ev_id, {"cofnij": True})
    sprawdz("cofnięcie odwołania przechodzi", kod == 200)
    conn = db.get_conn()
    w_grafiku = [e for e in cv.events_for_month(conn, "2026-12") if e["lead_id"] == l_odw]
    conn.close()
    sprawdz("po cofnięciu zajęcia wracają do grafiku", len(w_grafiku) == 1)

    # Kafel cyklu to WYSTĄPIENIE reguły — przycisk odwołania po `e.id` skasowałby
    # z grafiku cały pakiet. Lepiej nie dać przycisku, niż dać mylący.
    szablon = open("templates/kalendarz.html", encoding="utf-8").read()
    sprawdz("odwołanie w grafiku tylko dla wpisów niecyklicznych",
            # P31 wsunął przed ten warunek gałąź „to jest odwołane → przywróć",
            # więc `if` zamienił się w `elif`. Sam warunek jest ten sam i to
            # jego pilnujemy: przycisk odwołania nie ma prawa pojawić się na
            # kaflu cyklu.
            "{% elif e.typ not in TYPY_CYKLICZNE %}" in szablon
            and 'data-odwolaj="{{ e.id }}"' in szablon)

    # -----------------------------------------------------------------
    print("\nS19 — P30/P31: braki w DT widać w kalendarzu, odwołane mają swoją listę")

    # DT z kompletem danych i DT „zaczęty" — dokładnie to, co od P27 wolno
    # zapisać z terenu. Kalendarz jest jedynym miejscem, gdzie widać różnicę.
    conn = db.get_conn()
    p_id = conn.execute(
        "INSERT INTO placowki (nazwa, miejscowosc, typ) VALUES (?,?,?)",
        ("SP Braki", "01. Katowice", "01. Szkoła podstawowa")).lastrowid
    l_braki = conn.execute("INSERT INTO leady (placowka_id, handlowiec) VALUES (?,?)",
                           (p_id, "04. Chytry")).lastrowid
    conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od, trener, "
                 "ilosc_klas, ilosc_dzieci) VALUES (?,?,?,?,?,?,?)",
                 (l_braki, "DT", "2026-12-08", "09:00", "01. Małolepsza", 3, 60))
    id_niepelny = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, trener) VALUES (?,?,?,?)",
        (l_braki, "DT", "2026-12-09", "01. Małolepsza")).lastrowid
    conn.commit()

    evs = {e["id"]: e for e in cv.events_for_month(conn, "2026-12")}
    sprawdz("komplet danych = brak ostrzeżenia",
            evs[id_niepelny - 1]["braki"] == [], str(evs[id_niepelny - 1]["braki"]))
    sprawdz("DT bez godziny, klas i dzieci wymienia wszystkie trzy braki",
            evs[id_niepelny]["braki"] == ["godzina", "liczba klas", "liczba dzieci"],
            str(evs[id_niepelny]["braki"]))

    mac = cv.build_matrix(conn, "2026-12")
    sprawdz("licznik „do uzupełnienia” liczy tylko niepełne wpisy",
            mac["n_do_uzupelnienia"] == sum(1 for e in evs.values() if e["braki"]),
            "licznik %d" % mac["n_do_uzupelnienia"])
    tylko_braki = cv.build_matrix(conn, "2026-12", do_uzupelnienia=True)
    sprawdz("filtr braków zostawia same niepełne wpisy",
            tylko_braki["n_events"] == mac["n_do_uzupelnienia"]
            and tylko_braki["n_events"] < mac["n_events"],
            "%d z %d" % (tylko_braki["n_events"], mac["n_events"]))
    conn.close()

    # Zajęcia cykliczne nie mają liczby klas i nigdy nie będą miały — gdyby
    # wchodziły do licznika, „do uzupełnienia" pokazywałoby całą jesień.
    sprawdz("cykl nie jest „do uzupełnienia”",
            cv.braki_dt({"typ": "CYKLICZNE"}) == [])

    # P31 — lista odwołanych. Do 20.08 dało się je zobaczyć TYLKO na karcie
    # konkretnej szkoły, czyli trzeba było wiedzieć, której szukać.
    kod, _ = post("/api/event/%d/odwolaj" % ev_id, {"powod": "sala zajęta"})
    sprawdz("odwołanie na potrzeby listy przechodzi", kod == 200)
    conn = db.get_conn()
    lista = cv.events_for_month(conn, "2026-12", odwolane=True)
    grafik = cv.events_for_month(conn, "2026-12")
    conn.close()
    sprawdz("tryb „odwołane” pokazuje odwołane", [e["id"] for e in lista] == [ev_id],
            str([e["id"] for e in lista]))
    sprawdz("…i tylko je — grafik ich nie ma",
            ev_id not in [e["id"] for e in grafik])
    sprawdz("wpis niesie powód i osobę",
            (lista[0]["odwolanie"] or {}).get("powod") == "sala zajęta"
            and (lista[0]["odwolanie"] or {}).get("kto"),
            str(lista[0]["odwolanie"]))
    sprawdz("czynny wpis nie udaje odwołanego",
            grafik[0]["odwolanie"] is None)

    # Nazwa `odwolanie` jest celowo inna niż `odwolane`: to drugie w
    # calendar_view znaczy „odwołane WYSTĄPIENIE cyklu" i jest zerowane przy
    # każdym wpisie bez wyjątku. Przy pierwszym podejściu zjadło znacznik.
    zrodlo = open("calendar_view.py", encoding="utf-8").read()
    sprawdz("odwołanie spotkania ma własną nazwę pola",
            'e["odwolanie"] = ' in zrodlo and 'ev["odwolane"] = False' in zrodlo)

    szablon = open("templates/kalendarz.html", encoding="utf-8").read()
    sprawdz("wszystkie trzy widoki znaczą braki",
            szablon.count("tag tag-braki") == 3, str(szablon.count("tag tag-braki")))
    sprawdz("tryb odwołanych mówi wprost, że to nie grafik",
            "nie grafik" in szablon)
    sprawdz("z listy odwołanych da się przywrócić termin",
            szablon.count("btn-przywroc-event") == 3,
            str(szablon.count("btn-przywroc-event")))

    kod, _ = post("/api/event/%d/odwolaj" % ev_id, {"cofnij": True})
    sprawdz("sprzątanie po S19: termin wraca", kod == 200)

    # -----------------------------------------------------------------
    ok = sum(1 for _, w, _ in WYNIKI if w)
    zle = [n for n, w, _ in WYNIKI if not w]
    print("\n" + "=" * 62)
    print("WYNIK: %d/%d sprawdzeń OK" % (ok, len(WYNIKI)))
    if zle:
        print("NIEUDANE:")
        for n in zle:
            print("  -", n)
    print("=" * 62)
    return 0 if not zle else 1


if __name__ == "__main__":
    try:
        kod = main()
    finally:
        shutil.rmtree(TMP, ignore_errors=True)
    sys.exit(kod)
