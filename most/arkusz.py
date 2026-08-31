# -*- coding: utf-8 -*-
"""
Ten sam grafik co w `zajecia.json`, tylko do otwarcia dwuklikiem.

Po co drugi format tych samych danych: odbiorczyni powiedziała wprost „niech
zrobi skrypt, ja sobie na to zobaczę". Zanim ktokolwiek napisze import, ktoś musi
najpierw ZOBACZYĆ, co w tym pliku jest — i zdecydować, czy to jest to, o co
prosiła. JSON tego nie załatwia: żeby go przeczytać, trzeba już być programistą,
a wtedy rozmowa o zakresie odbywa się na końcu, a nie na początku.

Arkusz jest widokiem, nie źródłem. Buduje się WYŁĄCZNIE z gotowej migawki —
gdyby liczył cokolwiek sam, po pierwszej rozbieżności nikt by nie wiedział,
który plik kłamie.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ZIELONY = PatternFill("solid", fgColor="93C47D")     # kolor nagłówków z arkusza klienta
BURSZTYN = PatternFill("solid", fgColor="FFF2CC")    # wiersz z brakami
BOLD = Font(bold=True)

KOLUMNY_CYKLE = [
    ("id", "id", 26), ("stan", "stan", 10), ("title", "title", 26),
    ("school", "school", 42), ("address", "address", 26), ("city", "city", 18),
    ("region", "region", 18), ("weekday", "weekday", 8),
    ("start_time", "start_time", 10), ("end_time", "end_time", 10),
    ("starts_on", "starts_on", 12), ("ends_on", "ends_on", 12),
    ("every_n_weeks", "every_n_weeks", 8), ("liczba zajęć", "_ile", 8),
    ("odwołane terminy", "_ile_odwolanych", 8),
    ("trainer_name", "trainer_name", 20), ("printer_name", "printer_name", 20),
    ("room", "room", 8), ("group", "group", 8), ("equipment", "equipment", 22),
    ("classes", "classes", 8), ("children", "children", 8),
    ("salesperson", "salesperson", 18), ("braki", "_missing", 30),
]

KOLUMNY_DT = [
    ("id", "id", 26), ("stan", "stan", 10), ("status", "status", 10),
    ("title", "title", 40), ("school", "school", 42), ("address", "address", 26),
    ("city", "city", 18), ("region", "region", 18),
    ("date", "date", 12), ("start_time", "start_time", 10), ("end_time", "end_time", 10),
    ("starts_at", "starts_at", 20),
    ("trainer_name", "trainer_name", 20), ("printer_name", "printer_name", 20),
    ("room", "room", 8), ("classes", "classes", 8), ("children", "children", 8),
    ("salesperson", "salesperson", 18), ("braki", "_missing", 30),
]


def _wartosc(rec, klucz):
    if klucz == "_ile":
        return len(rec.get("occurrences") or [])
    if klucz == "_ile_odwolanych":
        return len(rec.get("cancelled_occurrences") or [])
    if klucz == "_missing":
        return ", ".join(rec.get("missing") or [])
    v = rec.get(klucz)
    return "" if v is None else v


def _arkusz(wb, tytul, kolumny, rekordy):
    ws = wb.create_sheet(tytul)
    for i, (etykieta, _k, szer) in enumerate(kolumny, 1):
        c = ws.cell(row=1, column=i, value=etykieta)
        c.font = BOLD
        c.fill = ZIELONY
        ws.column_dimensions[get_column_letter(i)].width = szer
    ws.freeze_panes = "A2"
    for w, rec in enumerate(rekordy, 2):
        braki = bool(rec.get("missing"))
        for i, (_e, klucz, _s) in enumerate(kolumny, 1):
            c = ws.cell(row=w, column=i, value=_wartosc(rec, klucz))
            # Rekord, którego odbiorca nie wstawi bez uzupełnienia, ma się rzucać
            # w oczy TU, a nie dopiero w jego logu importu.
            if braki:
                c.fill = BURSZTYN
    return ws


def _terminy(wb, cykle):
    """Jedno wystąpienie = jeden wiersz. To jest dowód, że nic nie gubimy."""
    ws = wb.create_sheet("Wystąpienia")
    for i, etykieta in enumerate(["id", "school", "data", "stan", "start_time",
                                  "end_time", "trainer_name"], 1):
        c = ws.cell(row=1, column=i, value=etykieta)
        c.font = BOLD
        c.fill = ZIELONY
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    w = 2
    for rec in cykle:
        pozycje = [(d, "zajęcia") for d in rec.get("occurrences") or []]
        pozycje += [(o["date"], "ODWOŁANE") for o in rec.get("cancelled_occurrences") or []]
        for data, stan in sorted(pozycje):
            for i, v in enumerate([rec["id"], rec.get("school"), data, stan,
                                   rec.get("start_time"), rec.get("end_time"),
                                   rec.get("trainer_name")], 1):
                ws.cell(row=w, column=i, value="" if v is None else v)
            w += 1
    return ws


def zbuduj(migawka):
    """Zwraca `bytes` gotowego pliku XLSX."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 52
    wiersze = [("format", migawka["format"]),
               ("wygenerowano", migawka["wygenerowano"]),
               ("profil", "%s (%s)" % (migawka["profil"], migawka["profil_etykieta"])),
               ("strefa czasu", migawka["strefa"])]
    wiersze += [(k, v) for k, v in migawka["liczby"].items()]
    for i, (k, v) in enumerate(wiersze, 1):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)

    _arkusz(wb, "Cykle", KOLUMNY_CYKLE, migawka["cykle"])
    _arkusz(wb, "DT", KOLUMNY_DT, migawka["dt"])
    _terminy(wb, migawka["cykle"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
