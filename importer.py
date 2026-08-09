# -*- coding: utf-8 -*-
"""
Import z plików klienta — bez ręcznego czyszczenia danych przed wgraniem.

Obsługiwane źródła:
  importuj_ph_nowy  — ich aktualny plik roboczy (`PH Nowy ... .xlsx`):
                      zakładka BAZA + 5 zakładek handlowców
  importuj_rspo     — czysty wyciąg z rejestru RSPO (szkoły/przedszkola/instytucje)
  wczytaj_demo      — jedno kliknięcie: realny arkusz + plansza STARTY jako dane pokazowe

TRZY DECYZJE, KTÓRE TU PODJĘTO (i dlaczego)

1. Kolumny mapujemy PO TREŚCI NAGŁÓWKA, nie po literze.
   W ich pliku ten sam sens ma różne litery: w `BAZA` kolumna Z to „Mail z wnioskiem
   o wynajem sali", a w `Zbiorczy` i u Olszewskiej Z to już „Dane do umowy WYPEŁNIA JULIA".
   Import po pozycjach przepisałby dane do złych pól.

2. Placówki DEDUPLIKUJEMY.
   Ta sama szkoła występuje w BAZIE (pełna nazwa z RSPO) i u handlowca (`MSP 1`, `sp1`).
   Klucz: numer RSPO gdy jest, inaczej znormalizowana nazwa krótka + miejscowość.
   Bez tego jedna szkoła byłaby kilkoma rekordami i kalendarz kłamałby.

3. Zajęcia cykliczne nie mają w ich pliku DATY PIERWSZYCH ZAJĘĆ — mają tylko dzień
   tygodnia i godzinę. Datę startu WYLICZAMY: pierwsze wystąpienie tego dnia tygodnia
   po dacie DT (a gdy DT nie ma — po 1 września roku szkolnego). Taki wpis jest
   oznaczany w uwagach, żeby koordynator mógł go poprawić, a nie żeby cicho zgadywać.
"""
import datetime as dt
import json
import os
import re

import openpyxl

import parsers as P
from db import alias_map, pl_fold, slownik_values, zapisz_log

# ---------------------------------------------------------------- konfiguracja

ZAKLADKI_HANDLOWCOW = ["Sacawa", "Olszewska", "Małolepsza", "Chytry", "Młynarczyk"]

# Zakładka z bazą placówek do rozdania. Nazwa JEST RUCHOMA: w pliku z czerwca
# nazywała się „BAZA", a w „PH PRÓBA Nowy dla handlowców.xlsx" (08.08.2026) już
# „Baza szkół Śląskie" — z 545 wierszami telefonów, maili i adresów. Sztywne
# porównanie do „BAZA" po cichu pomijało cały ten arkusz: import kończył się
# sukcesem, ale wchodziło 165 placówek z zakładek handlowców zamiast 545,
# czyli koordynatorka nie miała czego rozdawać. Dlatego dopasowujemy po
# POCZĄTKU nazwy — „baza…" łapie oba warianty i kolejne, które przyjdą.
PREFIKS_ZAKLADKI_BAZY = "baza"


def _zakladka_bazy(nazwy):
    """Pierwsza zakładka wyglądająca na bazę placówek albo None."""
    for n in nazwy:
        if pl_fold(n).startswith(PREFIKS_ZAKLADKI_BAZY):
            return n
    return None

# Nasze rodzaje słowników → rodzaje w `parsers` (parsers ma własne nazwy)
RODZAJ_PARSERS = {
    "handlowiec": "handlowiec",
    "trener": "trener",
    "miasto": "miejscowosc",
    "status_szkoly": "status_szkoly",
    "status_realizacji": "status_realizacji",
    "dt": "dt",
    "tak_nie": "tak_nie",
    "mail_dt": "mail_propozycja",
    "dzien_tyg": "dzien_tygodnia",
}

# Nagłówek w arkuszu klienta → nasze pole. Klucze porównujemy „na luźno"
# (bez ogonków, wielkości liter i nadmiarowych spacji), bo w jego pliku te same
# nagłówki mają różne końcówki i literówki.
MAPA_NAGLOWKOW = {
    "handlowiec": ("lead", "handlowiec"),
    "status szkoly": ("lead", "status_szkoly"),
    "status realizacji": ("lead", "status_realizacji"),
    "death line": ("lead", "deadline"),
    "deadline": ("lead", "deadline"),
    "ostateczny termin": ("lead", "deadline"),
    "miejscowosc": ("plac", "miejscowosc"),
    "numer placowki": ("plac", "nazwa"),
    "nazwa placowki": ("plac", "nazwa"),
    "adres placowki": ("plac", "adres"),
    "osoby decyzyjne i kontakt": ("plac", "osoba_kontakt"),
    "numer telefonu": ("plac", "telefon"),
    "telefon": ("plac", "telefon"),
    "mail": ("plac", "mail"),
    "uwagi": ("lead", "uwagi"),
    "dt": ("lead", "dt"),
    "data dt": ("ev_dt", "data"),
    "godzina dt": ("ev_dt", "godz"),
    "prowadzacy dt": ("ev_dt", "trener"),
    "numer sali dt": ("ev_dt", "numer_sali"),
    "mail propozycja lub ustalenie dt": ("lead", "mail_dt"),
    "ilosc klas 1-4": ("ev_dt", "ilosc_klas"),
    "ilosc klas": ("ev_dt", "ilosc_klas"),
    "ilosc dzieci w klasach": ("ev_dt", "ilosc_dzieci"),
    "ilosc dzieci": ("ev_dt", "ilosc_dzieci"),
    "mail do rodzicow na dziennik elektroniczny": ("lead", "mail_rodzice"),
    "cykle": ("lead", "cykle"),
    "zajecia cykliczne (dzien tygodnia)": ("ev_cy", "cykl_dzien"),
    "numer sali cykle": ("ev_cy", "numer_sali"),
    "zajecia cykliczne (godzina)": ("ev_cy", "godz"),
    "zajecia cykliczne (sala komputerowa/chromebooki)": ("ev_cy", "sprzet"),
    "trener": ("ev_cy", "trener"),
    "mail z wnioskiem o wynajem sali": ("lead", "mail_wynajem"),
    "dane do umowy wypelnia julia": ("lead", "julia_dane_umowy"),
    "standardy ochrony maloletnich wypelnia julia": ("lead", "julia_standardy"),
    "oswiadczenia trenerow do standardow wypelnia julia": ("lead", "julia_oswiadczenia"),
    "zaswiadczenie o niekaralnosci wypelnia julia": ("lead", "julia_niekaralnosc"),
    "podanie o wynajem sali wypelnia julia": ("lead", "julia_podanie_sala"),
    "umowa podpisana wypelnia julia": ("lead", "julia_umowa"),
    "librus wypelnia julia": ("lead", "julia_librus"),
    "nr rspo": ("plac", "rspo"),
    "rspo": ("plac", "rspo"),
    "typ": ("plac", "typ"),
}

# pola leada, które trzymają wartość słownikową (do normalizacji)
POLA_SLOWNIKOWE_LEAD = {
    "handlowiec": "handlowiec",
    "status_szkoly": "status_szkoly",
    "status_realizacji": "status_realizacji",
    "dt": "dt",
    "cykle": "tak_nie",
    "mail_dt": "mail_dt",
    "mail_rodzice": "tak_nie",
    "mail_wynajem": "tak_nie",
    "julia_dane_umowy": "tak_nie",
    "julia_standardy": "tak_nie",
    "julia_oswiadczenia": "tak_nie",
    "julia_niekaralnosc": "tak_nie",
    "julia_podanie_sala": "tak_nie",
    "julia_umowa": "tak_nie",
    "julia_librus": "tak_nie",
}

TYP_PLACOWKI_SLOWNIK = {
    "szkoła": "01. Szkoła podstawowa",
    "przedszkole": "02. Przedszkole",
    "instytucja kultury": "03. Instytucja kultury",
    "nieznany": "04. Inna",
}


def _fold_naglowek(s):
    """Nagłówek → klucz porównawczy: bez ogonków, bez wielokrotnych spacji, lowercase."""
    if s is None:
        return ""
    s = str(s)
    for a, b in (("ą", "a"), ("ć", "c"), ("ę", "e"), ("ł", "l"), ("ń", "n"),
                 ("ó", "o"), ("ś", "s"), ("ź", "z"), ("ż", "z"),
                 ("Ą", "a"), ("Ć", "c"), ("Ę", "e"), ("Ł", "l"), ("Ń", "n"),
                 ("Ó", "o"), ("Ś", "s"), ("Ź", "z"), ("Ż", "z")):
        s = s.replace(a, b)
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


# ---------------------------------------------------------------- normalizacja

class Normalizator:
    """
    Normalizacja wartości słownikowych: najpierw `parsers` (zna literówki i rozjazdy
    numeracji), potem kontrola względem SŁOWNIKA W BAZIE i aliasów z bazy.
    Wszystko, czego nie dało się przypisać, ląduje w raporcie — nic nie ginie po cichu.
    """

    def __init__(self, conn):
        self.slowniki = {r: set(slownik_values(conn, r)) for r in RODZAJ_PARSERS}
        self.slowniki["typ_placowki"] = set(slownik_values(conn, "typ_placowki"))
        self.slowniki["sprzet"] = set(slownik_values(conn, "sprzet"))
        self.aliasy_db = alias_map(conn)
        # aliasy z bazy przemapowane na nazwy rodzajów używane przez `parsers`
        self.aliasy_dla_parsers = {}
        for moj, ich in RODZAJ_PARSERS.items():
            if moj in self.aliasy_db:
                self.aliasy_dla_parsers.setdefault(ich, {}).update(self.aliasy_db[moj])
        for ich, mapa in P.ALIASY.items():
            self.aliasy_dla_parsers.setdefault(ich, {})
            for k, v in mapa.items():
                self.aliasy_dla_parsers[ich].setdefault(k, v)
        self.nierozpoznane = {}

    def __call__(self, wartosc, rodzaj):
        if wartosc is None or str(wartosc).strip() == "":
            return None
        surowa = str(wartosc).strip()
        dozwolone = self.slowniki.get(rodzaj, set())

        # 1) aliasy z bazy (nasze, pełne — m.in. nazwy z planszy STARTY)
        kandydat = self.aliasy_db.get(rodzaj, {}).get(surowa)

        # 2) parsers: literówki, rozjazdy numeracji, dopasowanie rozmyte
        if kandydat is None and rodzaj in RODZAJ_PARSERS:
            kandydat = P.norm_slownik(surowa, RODZAJ_PARSERS[rodzaj],
                                      aliasy=self.aliasy_dla_parsers)
        if kandydat is None:
            kandydat = surowa

        if kandydat in dozwolone:
            return kandydat
        # 3) jeszcze raz aliasy — parsers mógł zwrócić wariant, który mamy w aliasach
        drugie = self.aliasy_db.get(rodzaj, {}).get(kandydat)
        if drugie and drugie in dozwolone:
            return drugie
        # 4) dopasowanie po części nazwowej do NASZEGO słownika
        klucz = _fold_naglowek(P.strip_prefix(kandydat)[1])
        for d in dozwolone:
            if _fold_naglowek(P.strip_prefix(d)[1]) == klucz:
                return d

        self.nierozpoznane.setdefault(rodzaj, {})
        self.nierozpoznane[rodzaj][surowa] = self.nierozpoznane[rodzaj].get(surowa, 0) + 1
        return kandydat


# ---------------------------------------------------------------- arkusz → wiersze

def _mapa_kolumn(ws, wiersz_naglowka=1):
    """{indeks_kolumny: (grupa, pole)} na podstawie treści nagłówków."""
    mapa = {}
    for c in range(1, (ws.max_column or 1) + 1):
        h = _fold_naglowek(ws.cell(row=wiersz_naglowka, column=c).value)
        if not h:
            continue
        if h in MAPA_NAGLOWKOW:
            mapa[c] = MAPA_NAGLOWKOW[h]
    return mapa


def _pierwszy_wiersz_danych(ws, mapa):
    """
    W ich pliku dane zaczynają się w 4. wierszu (wiersze 2–3 są puste/ozdobne),
    ale w widokach już w 2. Zamiast zgadywać — szukamy pierwszego wiersza,
    w którym cokolwiek jest w kolumnie nazwy placówki lub handlowca.
    """
    kol_nazwa = [c for c, (g, p) in mapa.items() if (g, p) == ("plac", "nazwa")]
    kol_h = [c for c, (g, p) in mapa.items() if (g, p) == ("lead", "handlowiec")]
    kolumny = kol_nazwa + kol_h
    if not kolumny:
        return 2
    for r in range(2, min((ws.max_row or 2), 40) + 1):
        for c in kolumny:
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                return r
    return 2


def _czytaj_arkusz(ws):
    """Zwraca listę słowników {grupa: {pole: wartosc}} — jeden na wiersz z danymi."""
    mapa = _mapa_kolumn(ws)
    if not mapa:
        return [], {}
    start = _pierwszy_wiersz_danych(ws, mapa)
    wiersze = []
    for r in range(start, (ws.max_row or start) + 1):
        rec = {"lead": {}, "plac": {}, "ev_dt": {}, "ev_cy": {}}
        pusty = True
        for c, (grupa, pole) in mapa.items():
            v = ws.cell(row=r, column=c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            rec[grupa][pole] = v
            pusty = False
        if pusty:
            continue
        # wiersz bez nazwy placówki jest bezużyteczny — nie da się go z niczym powiązać
        if not str(rec["plac"].get("nazwa") or "").strip():
            continue
        rec["_wiersz"] = r
        wiersze.append(rec)
    return wiersze, mapa


# ---------------------------------------------------------------- zapis do bazy

def _klucz_placowki(nazwa_krotka, miejscowosc, rspo):
    if rspo:
        return ("rspo", str(rspo).strip())
    return ("nm", _fold_naglowek(nazwa_krotka), _fold_naglowek(miejscowosc or ""))


class Zapisywacz:
    def __init__(self, conn, norm, zrodlo):
        self.conn = conn
        self.norm = norm
        self.zrodlo = zrodlo
        self.indeks = {}
        self.raport = {"placowki": 0, "leady": 0, "eventy": 0, "pominiete": 0,
                       "placowki_scalone": 0, "cykle_z_wyliczona_data": 0,
                       "nierozpoznane": {}, "uwagi": []}
        self._zbuduj_indeks()

    def _zbuduj_indeks(self):
        for r in self.conn.execute("SELECT id, nazwa, miejscowosc, rspo FROM placowki"):
            _, krotka = P.norm_placowka(r["nazwa"])
            self.indeks[_klucz_placowki(krotka, r["miejscowosc"], r["rspo"])] = r["id"]

    def placowka(self, dane):
        nazwa_pelna = str(dane.get("nazwa") or "").strip()
        if not nazwa_pelna:
            return None
        typ_txt, krotka = P.norm_placowka(nazwa_pelna)
        miejscowosc = self.norm(dane.get("miejscowosc"), "miasto")
        rspo = dane.get("rspo")
        rspo = str(rspo).strip() if rspo not in (None, "") else None
        klucz = _klucz_placowki(krotka, miejscowosc, rspo)

        if klucz in self.indeks:
            pid = self.indeks[klucz]
            self.raport["placowki_scalone"] += 1
            # uzupełniamy braki, ale nie nadpisujemy tego, co już jest
            self._uzupelnij_placowke(pid, dane, miejscowosc, typ_txt, nazwa_pelna)
            return pid

        typ = self.norm(dane.get("typ"), "typ_placowki") or \
            TYP_PLACOWKI_SLOWNIK.get(typ_txt, "04. Inna")
        cur = self.conn.execute(
            "INSERT INTO placowki (rspo, nazwa, typ, miejscowosc, adres, osoba_kontakt, "
            "telefon, mail, zrodlo) VALUES (?,?,?,?,?,?,?,?,?)",
            (rspo, nazwa_pelna, typ, miejscowosc,
             _txt(dane.get("adres")), _txt(dane.get("osoba_kontakt")),
             P.parse_phone(dane.get("telefon")), _txt(dane.get("mail")), self.zrodlo))
        pid = cur.lastrowid
        self.indeks[klucz] = pid
        self.raport["placowki"] += 1
        return pid

    def _uzupelnij_placowke(self, pid, dane, miejscowosc, typ_txt, nazwa_pelna):
        stary = self.conn.execute("SELECT * FROM placowki WHERE id=?", (pid,)).fetchone()
        upd, par = [], []
        kandydaci = {
            "adres": _txt(dane.get("adres")),
            "osoba_kontakt": _txt(dane.get("osoba_kontakt")),
            "telefon": P.parse_phone(dane.get("telefon")),
            "mail": _txt(dane.get("mail")),
            "miejscowosc": miejscowosc,
        }
        # dłuższa nazwa jest zwykle pełną nazwą z RSPO — lepsza niż „sp1"
        if len(nazwa_pelna) > len(stary["nazwa"] or ""):
            kandydaci["nazwa"] = nazwa_pelna
        for k, v in kandydaci.items():
            if v and not (stary[k] or "").strip():
                upd.append("%s=?" % k)
                par.append(v)
            elif k == "nazwa" and v:
                upd.append("nazwa=?")
                par.append(v)
        if upd:
            par.append(pid)
            self.conn.execute("UPDATE placowki SET %s, updated_at=datetime('now') "
                              "WHERE id=?" % ", ".join(upd), par)

    def lead(self, placowka_id, dane, domyslny_handlowiec=None):
        pola = {}
        for pole, wartosc in dane.items():
            if pole in POLA_SLOWNIKOWE_LEAD:
                pola[pole] = self.norm(wartosc, POLA_SLOWNIKOWE_LEAD[pole])
            elif pole == "deadline":
                d = P.parse_date(wartosc)
                pola["deadline"] = d.isoformat() if d else None
            elif pole == "uwagi":
                pola["uwagi"] = _txt(wartosc)
        if not pola.get("handlowiec") and domyslny_handlowiec:
            pola["handlowiec"] = domyslny_handlowiec
        if not pola.get("status_realizacji"):
            pola["status_realizacji"] = ("00. Nieprzydzielony"
                                         if not pola.get("handlowiec")
                                         else "01. Próba kontaktu (Brak konkretów)")

        # czy ta placówka ma już leada? (ta sama szkoła w BAZIE i u handlowca)
        istn = self.conn.execute(
            "SELECT id, handlowiec, status_realizacji, deadline FROM leady "
            "WHERE placowka_id=?", (placowka_id,)).fetchone()
        if istn:
            upd, par = [], []
            for k, v in pola.items():
                if v is None:
                    continue
                if k == "status_realizacji":
                    # nie cofamy procesu: 03. > 02b. > 02. > 01. > 00.
                    if _ranga(v) <= _ranga(istn["status_realizacji"]):
                        continue
                elif (istn[k] if k in istn.keys() else None):
                    continue
                upd.append("%s=?" % k)
                par.append(v)
            if upd:
                par.append(istn["id"])
                self.conn.execute("UPDATE leady SET %s, updated_at=datetime('now') "
                                  "WHERE id=?" % ", ".join(upd), par)
            return istn["id"]

        kolumny = ["placowka_id"] + list(pola.keys())
        cur = self.conn.execute(
            "INSERT INTO leady (%s) VALUES (%s)"
            % (", ".join(kolumny), ", ".join(["?"] * len(kolumny))),
            [placowka_id] + list(pola.values()))
        self.raport["leady"] += 1
        return cur.lastrowid

    def event_dt(self, lead_id, dane):
        data = P.parse_date(dane.get("data"))
        trener = self.norm(dane.get("trener"), "trener")
        if not data and not trener:
            return None
        od, do = None, None
        if dane.get("godz") is not None:
            t = P.parse_time(dane.get("godz"))
            if t:
                od = t.strftime("%H:%M")
            else:
                a, b = P.parse_time_range(dane.get("godz"))
                od = a.strftime("%H:%M") if a else None
                do = b.strftime("%H:%M") if b else None
        return self._wstaw_event(lead_id, {
            "typ": "DT",
            "data": data.isoformat() if data else None,
            "godz_od": od, "godz_do": do, "trener": trener,
            "numer_sali": _txt(dane.get("numer_sali")),
            "ilosc_klas": P.parse_int_loose(dane.get("ilosc_klas")),
            "ilosc_dzieci": P.parse_int_loose(dane.get("ilosc_dzieci")),
        })

    def event_cykl(self, lead_id, dane, data_dt=None):
        """
        Zajęcia cykliczne. W arkuszu klienta NIE MA daty pierwszych zajęć —
        wyliczamy ją z dnia tygodnia. Każdy taki wpis jest oznaczony w uwagach.
        """
        dni = P.parse_dni_tygodnia(dane.get("cykl_dzien"))
        trener = self.norm(dane.get("trener"), "trener")
        godz = dane.get("godz")
        if not dni and not trener and not godz:
            return None
        od, do = None, None
        if godz is not None:
            a, b = P.parse_time_range(godz)
            od = a.strftime("%H:%M") if a else None
            do = b.strftime("%H:%M") if b else None
            if not od:
                t = P.parse_time(godz)
                od = t.strftime("%H:%M") if t else None
        sprzet = self.norm(dane.get("sprzet"), "sprzet")
        ids = []
        for dzien in (dni or [None]):
            data = None
            if dzien:
                data = _pierwsza_data_dnia(dzien, data_dt)
                self.raport["cykle_z_wyliczona_data"] += 1
            ids.append(self._wstaw_event(lead_id, {
                "typ": "CYKLICZNE",
                "data": data.isoformat() if data else None,
                "godz_od": od, "godz_do": do, "trener": trener,
                "cykl_dzien": dzien, "sprzet": sprzet,
                "numer_sali": _txt(dane.get("numer_sali")),
                "uwagi": "data pierwszych zajęć wyliczona z dnia tygodnia — do potwierdzenia"
                         if data else None,
            }))
        return [i for i in ids if i]

    def _wstaw_event(self, lead_id, pola):
        pola = {k: v for k, v in pola.items() if v not in (None, "")}
        if len(pola) <= 1:
            return None
        pola.setdefault("typ", "DT")
        kolumny = ["lead_id"] + list(pola.keys())
        cur = self.conn.execute(
            "INSERT INTO eventy (%s) VALUES (%s)"
            % (", ".join(kolumny), ", ".join(["?"] * len(kolumny))),
            [lead_id] + list(pola.values()))
        self.raport["eventy"] += 1
        return cur.lastrowid


_RANGI = {"00.": 0, "01.": 1, "02.": 2, "02b": 3, "03.": 5, "04.": 4}


def _ranga(status):
    if not status:
        return -1
    return _RANGI.get(str(status)[:3], 0)


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    m = re.match(r'^\s*=\s*"(.*)"\s*$', s)   # ="601290441"
    if m:
        s = m.group(1).strip()
    return s or None


# W planszy STARTY pola „ZASTĘPSTWO:" i „DRUKARZ:" bywają wypełnione znakami
# zamiast nazwiskiem (252 z 286 wpisów ma tam samo ':'). To nie osoba — to puste pole.
_SMIECI_OSOBA = {":", "-", "?", "!", "!!!", "!!!!", "!!!!!", "!!!!!!!!!",
                 "*osoba szkolona:", "brak", "brak!!!", "xxx"}


def _osoba_raw(v):
    """Zapis osoby z arkusza → tekst do normalizacji, albo None gdy to śmieć."""
    s = (str(v) if v is not None else "").strip()
    if not s or s.strip(":!?*- ") == "":
        return None
    if s.lower() in _SMIECI_OSOBA:
        return None
    return s.strip(": ").strip() or None


_DNI_IDX = {"poniedziałek": 0, "wtorek": 1, "środa": 2, "czwartek": 3,
            "piątek": 4, "sobota": 5, "niedziela": 6}


def _pierwsza_data_dnia(dzien, po_dacie=None):
    """Pierwsze wystąpienie danego dnia tygodnia po `po_dacie` (domyślnie 1 września)."""
    idx = _DNI_IDX.get(dzien)
    if idx is None:
        return None
    baza = po_dacie
    if isinstance(baza, str):
        baza = P.parse_date(baza)
    if not baza:
        dzisiaj = dt.date.today()
        rok = dzisiaj.year if dzisiaj.month >= 9 else dzisiaj.year - 1
        baza = dt.date(rok, 9, 1)
    delta = (idx - baza.weekday()) % 7
    return baza + dt.timedelta(days=delta or 7)


# ---------------------------------------------------------------- import: PH Nowy

def importuj_ph_nowy(conn, sciezka, replace=False):
    """
    Import ich pliku roboczego. Kolejność ma znaczenie:
    najpierw BAZA (pełne nazwy z RSPO, adresy, telefony), potem zakładki handlowców
    (statusy, terminy, DT, cykle) — dzięki temu placówka ma pełne dane, a lead pełny proces.
    """
    if replace:
        wyczysc(conn)
    wb = openpyxl.load_workbook(sciezka, data_only=True, read_only=False)
    norm = Normalizator(conn)
    zap = Zapisywacz(conn, norm, "arkusz:PH Nowy")

    baza = _zakladka_bazy(wb.sheetnames)
    kolejnosc = ([baza] if baza else []) + \
                [z for z in ZAKLADKI_HANDLOWCOW if z in wb.sheetnames]
    if not baza:
        # Widoczne w raporcie, bo to różnica między „mamy co rozdawać"
        # a „koordynatorka patrzy na pustą listę".
        zap.raport["uwagi"].append(
            "UWAGA: nie znaleziono zakładki z bazą placówek (nazwa zaczynająca się "
            "od „Baza…”) — weszły tylko szkoły z zakładek handlowców")
    if not kolejnosc:
        # nieznany plik — bierzemy pierwszą zakładkę, która ma rozpoznawalne nagłówki
        kolejnosc = [n for n in wb.sheetnames if _mapa_kolumn(wb[n])][:1]

    for nazwa_zakladki in kolejnosc:
        ws = wb[nazwa_zakladki]
        wiersze, mapa = _czytaj_arkusz(ws)
        if not wiersze:
            zap.raport["uwagi"].append(
                "zakładka %r: brak rozpoznanych wierszy" % nazwa_zakladki)
            continue
        domyslny_h = None
        if nazwa_zakladki in ZAKLADKI_HANDLOWCOW:
            domyslny_h = norm(nazwa_zakladki, "handlowiec")
        for rec in wiersze:
            pid = zap.placowka(rec["plac"])
            if not pid:
                zap.raport["pominiete"] += 1
                continue
            lid = zap.lead(pid, rec["lead"], domyslny_handlowiec=domyslny_h)
            zap.event_dt(lid, rec["ev_dt"])
            if rec["ev_cy"]:
                zap.event_cykl(lid, rec["ev_cy"], data_dt=rec["ev_dt"].get("data"))
        zap.raport["uwagi"].append(
            "zakładka %r: %d wierszy, kolumn rozpoznanych %d"
            % (nazwa_zakladki, len(wiersze), len(mapa)))

    conn.commit()
    zap.raport["nierozpoznane"] = norm.nierozpoznane
    wb.close()
    return zap.raport


# ---------------------------------------------------------------- import: RSPO

NAGLOWKI_RSPO = {
    "rspo": "rspo", "numer rspo": "rspo", "nr rspo": "rspo",
    "nazwa": "nazwa", "nazwa placowki": "nazwa", "nazwa szkoly": "nazwa",
    "typ": "typ", "typ placowki": "typ", "rodzaj placowki": "typ",
    "miejscowosc": "miejscowosc", "miasto": "miejscowosc", "gmina": "miejscowosc",
    "adres": "adres", "ulica": "adres", "adres placowki": "adres",
    "telefon": "telefon", "numer telefonu": "telefon",
    "mail": "mail", "email": "mail", "e-mail": "mail",
    "dyrektor": "osoba_kontakt", "osoba kontaktowa": "osoba_kontakt",
}


def importuj_rspo(conn, sciezka, replace=False, arkusz=None):
    """
    Import czystej bazy z rejestru RSPO. Wgrywa TYLKO placówki + leada w stanie
    „nieprzydzielony" — przypisaniem zajmuje się koordynator na ekranie /baza.
    """
    if replace:
        wyczysc(conn, tylko_nieprzydzielone=True)
    wb = openpyxl.load_workbook(sciezka, data_only=True)
    ws = wb[arkusz] if arkusz and arkusz in wb.sheetnames else wb[wb.sheetnames[0]]
    norm = Normalizator(conn)
    zap = Zapisywacz(conn, norm, "rspo")

    # nagłówek: pierwszy wiersz, w którym rozpoznamy min. 2 kolumny
    wiersz_h, mapa = None, {}
    for r in range(1, min((ws.max_row or 1), 10) + 1):
        m = {}
        for c in range(1, (ws.max_column or 1) + 1):
            h = _fold_naglowek(ws.cell(row=r, column=c).value)
            if h in NAGLOWKI_RSPO:
                m[c] = NAGLOWKI_RSPO[h]
        if len(m) >= 2:
            wiersz_h, mapa = r, m
            break
    if not mapa:
        wb.close()
        return {"placowki": 0, "leady": 0, "eventy": 0, "pominiete": 0,
                "uwagi": ["nie rozpoznano nagłówków — oczekiwane m.in.: "
                          "nazwa, miejscowość, adres, telefon, mail, RSPO"],
                "nierozpoznane": {}}

    for r in range(wiersz_h + 1, (ws.max_row or wiersz_h) + 1):
        dane = {}
        for c, pole in mapa.items():
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                dane[pole] = v
        if not dane.get("nazwa"):
            continue
        pid = zap.placowka(dane)
        if pid:
            zap.lead(pid, {})
    conn.commit()
    zap.raport["nierozpoznane"] = norm.nierozpoznane
    zap.raport["uwagi"].append("nagłówek w wierszu %d, kolumn rozpoznanych %d"
                               % (wiersz_h, len(mapa)))
    wb.close()
    return zap.raport


# ---------------------------------------------------------------- import: STARTY (demo)

def importuj_starty_json(conn, sciezka, rok=None, miesiac=6):
    """
    Wczytanie planszy STARTY (znormalizowane dane z zakładki „STARTY CZERWIEC").
    Służy do pokazania widoku Starty na realnym wolumenie (~30 zajęć dziennie),
    bo w bieżącym pliku klienta cyklicznych jeszcze prawie nie ma.

    UWAGA NA DUBLE. Plansza w arkuszu to SZABLON TYGODNIOWY kopiowany w dół:
    w czerwcu tydzień 2 ma 149 z 155 komórek identycznych z tygodniem 1.
    Gdybyśmy wczytali każdy wpis jako osobne zajęcia cykliczne, każda grupa
    powtarzałaby się dwa razy w tygodniu — i kalendarz pokazałby dwa razy więcej
    zajęć, niż firma realnie prowadzi (a licznik kolizji oszalałby).

    Dlatego: pierwsze wystąpienie grupy = REGUŁA cyklu, a różnice w kolejnych
    tygodniach (inny trener, zastępstwo) = WYJĄTEK na konkretną datę.
    Dokładnie tak, jak klient realnie pracuje.
    """
    if not os.path.exists(sciezka):
        return {"placowki": 0, "leady": 0, "eventy": 0, "pominiete": 0,
                "uwagi": ["brak pliku %s" % sciezka], "nierozpoznane": {}}
    dane = json.load(open(sciezka, encoding="utf-8"))
    norm = Normalizator(conn)
    zap = Zapisywacz(conn, norm, "starty")
    rok = rok or dt.date.today().year
    zap.raport["wyjatki"] = 0
    zap.raport["duble_pominiete"] = 0
    znane_grupy = {}      # klucz grupy → (event_id, data_pierwszych, trener)

    dane = sorted(dane, key=lambda x: (x.get("tydzien") or 0, x.get("dzien_nr") or 0))
    for w in dane:
        dzien_nr = w.get("dzien_nr")
        if not dzien_nr:
            continue
        try:
            data = dt.date(rok, miesiac, int(dzien_nr))
        except ValueError:
            continue
        nazwa = (w.get("placowka") or "").strip()
        if not nazwa:
            continue
        # w STARTY miejscowość jest wtopiona w nazwę/adres — wyciągamy ją zgrubnie
        miejscowosc = _miasto_ze_startow(nazwa, w.get("adres") or "")
        pid = zap.placowka({"nazwa": nazwa, "adres": w.get("adres"),
                            "miejscowosc": miejscowosc})
        if not pid:
            zap.raport["pominiete"] += 1
            continue
        # Bierzemy ZAPIS Z ARKUSZA (`trener_raw`), nie wstępnie znormalizowaną nazwę.
        # Jedna ścieżka normalizacji (słownik + aliasy w bazie) zamiast dwóch, które
        # się rozjeżdżają: aliasy znają 'GAWRON KORNELIA', nie znają 'Kornelia Gawron'.
        raw = w.get("trener_raw") or ""
        czesci = [_osoba_raw(c) for c in raw.split("+")]
        czesci = [c for c in czesci if c] or [None]
        trener = norm(czesci[0], "trener") if czesci[0] else None
        trener2 = norm(czesci[1], "trener") if len(czesci) > 1 else None
        zast = _osoba_raw(w.get("zastepstwo_raw"))
        dru = _osoba_raw(w.get("drukarz_raw"))
        # UWAGA: plansza STARTY to grafik ZAJĘĆ CYKLICZNYCH, nie DT.
        # „START" oznacza tam pierwsze zajęcia nowej grupy (inauguracja) — termin
        # jednorazowy, więc nie rozwijamy go tygodniowo. Mapowanie START→DT byłoby błędem.
        typ_ev = "START" if w.get("typ") == "START" else "CYKLICZNE"
        zast_n = norm(zast, "trener") if zast else None
        dru_n = norm(dru, "trener") if dru else None
        dzien_n = norm(w.get("dzien_tyg"), "dzien_tyg")
        lid = zap.lead(pid, {"status_realizacji": "03b. Grupa cykliczna otwarta",
                             "cykle": "01. Tak"})

        # klucz tożsamości grupy cyklicznej — bez daty i bez trenera,
        # bo trener może się w danym tygodniu zmienić (i to jest właśnie wyjątek)
        klucz = (pid, (w.get("grupa") or ""), dzien_n or "",
                 w.get("godz_od") or "", typ_ev)

        if typ_ev == "CYKLICZNE" and klucz in znane_grupy:
            eid, trener_reguly = znane_grupy[klucz]
            rozne_trener = trener and trener_reguly and trener != trener_reguly
            if rozne_trener or zast_n:
                conn.execute(
                    "INSERT OR REPLACE INTO wyjatki_cyklu "
                    "(event_id, data, trener, zastepstwo, uwagi) VALUES (?,?,?,?,?)",
                    (eid, data.isoformat(),
                     trener if rozne_trener else None, zast_n,
                     "z planszy STARTY: %s" % ("zmiana prowadzącego" if rozne_trener
                                               else "zastępstwo")))
                zap.raport["wyjatki"] += 1
            else:
                zap.raport["duble_pominiete"] += 1
            continue

        eid = zap._wstaw_event(lid, {
            "typ": typ_ev,
            "data": data.isoformat(),
            "co_ile_tygodni": 1 if typ_ev == "CYKLICZNE" else None,
            "godz_od": w.get("godz_od"), "godz_do": w.get("godz_do"),
            "trener": trener, "trener2": trener2,
            "zastepstwo": zast_n,
            "drukarz": dru_n,
            "grupa": w.get("grupa") or None,
            "sprzet": norm(w.get("sprzet"), "sprzet"),
            "cykl_dzien": dzien_n,
            "kod_tinkercad": w.get("kod_tinkercad") or None,
            "link_tinkercad": w.get("link_tinkercad") or None,
        })
        if eid and typ_ev == "CYKLICZNE":
            znane_grupy[klucz] = (eid, trener)
    conn.commit()
    zap.raport["nierozpoznane"] = norm.nierozpoznane
    wb_uwaga = ("plansza STARTY wczytana jako %d.%02d — to dane pokazowe "
                "z zeszłego roku szkolnego" % (rok, miesiac))
    zap.raport["uwagi"].append(wb_uwaga)
    return zap.raport


_MIASTA_STARTY = [
    "Strzyżowice", "Piekary Śląskie", "Siemianowice Śląskie", "Dąbrowa Górnicza",
    "Sosnowiec", "Katowice", "Chorzów", "Będzin", "Knurów", "Rybnik", "Tychy",
    "Mikołów", "Orzesze", "Żory", "Zabrze", "Ruda Śląska", "Świętochłowice",
    "Jaworzno", "Pszczyna", "Łaziska Górne", "Ornontowice", "Wyry", "Gostyń",
]


def _miasto_ze_startow(nazwa, adres):
    """W planszy STARTY miasto jest wtopione w nazwę lub adres — wyłuskujemy je."""
    tekst = "%s %s" % (nazwa, adres)
    low = _fold_naglowek(tekst)
    for m in _MIASTA_STARTY:
        if _fold_naglowek(m) in low:
            return m
    return None


# ---------------------------------------------------------------- demo / czyszczenie

SCIEZKA_PH_NOWY = os.environ.get(
    "PLIK_PH_NOWY",
    r"C:\XEN\AI-szkolenie\LIPIEC2026\PH Nowy  Nad którym pracuję jako główny  .xlsx")
SCIEZKA_STARTY = os.path.join(os.path.dirname(__file__), "docs", "design",
                              "starty_normalized.json")


def wczytaj_demo(conn, ph_nowy=None, starty=None):
    """Jedno kliknięcie: realny arkusz klienta + plansza STARTY."""
    raport = {"placowki": 0, "leady": 0, "eventy": 0, "pominiete": 0,
              "uwagi": [], "nierozpoznane": {}, "cykle_z_wyliczona_data": 0,
              "placowki_scalone": 0}
    sciezka = ph_nowy or SCIEZKA_PH_NOWY
    if os.path.exists(sciezka):
        r1 = importuj_ph_nowy(conn, sciezka)
        _scal_raport(raport, r1)
    else:
        raport["uwagi"].append("nie znaleziono pliku %s" % sciezka)
    sc_st = starty or SCIEZKA_STARTY
    if os.path.exists(sc_st):
        r2 = importuj_starty_json(conn, sc_st)
        _scal_raport(raport, r2)
    return raport


def _scal_raport(a, b):
    for k in ("placowki", "leady", "eventy", "pominiete", "placowki_scalone",
              "cykle_z_wyliczona_data", "wyjatki", "duble_pominiete"):
        a[k] = a.get(k, 0) + b.get(k, 0)
    a["uwagi"] += b.get("uwagi", [])
    for rodzaj, mapa in (b.get("nierozpoznane") or {}).items():
        a["nierozpoznane"].setdefault(rodzaj, {})
        for w, n in mapa.items():
            a["nierozpoznane"][rodzaj][w] = a["nierozpoznane"][rodzaj].get(w, 0) + n
    return a


def wyczysc(conn, tylko_nieprzydzielone=False):
    """Czyszczenie danych operacyjnych. Słowniki i aliasy zostają."""
    if tylko_nieprzydzielone:
        conn.execute("DELETE FROM eventy WHERE lead_id IN "
                     "(SELECT id FROM leady WHERE handlowiec IS NULL OR handlowiec='')")
        conn.execute("DELETE FROM leady WHERE handlowiec IS NULL OR handlowiec=''")
        conn.execute("DELETE FROM placowki WHERE id NOT IN "
                     "(SELECT placowka_id FROM leady)")
    else:
        conn.executescript("DELETE FROM eventy; DELETE FROM leady; "
                           "DELETE FROM placowki; DELETE FROM log;")
    conn.commit()
