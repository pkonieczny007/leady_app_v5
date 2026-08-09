# -*- coding: utf-8 -*-
"""
Rejony trenerów z zakładki „Trenerzy regiony" → tabela `rejony`.

PO CO
Rejon („kto po jakich miastach jeździ") jest jednym z czterech składników
podpowiedzi trenera w formularzu — `przydzial.kandydaci` podbija nim kolejność
i wypisuje „jeździ tu". Sprawdzone 09.08: tabela `rejony` była **pusta**
(0 z 40 trenerów), choć klient ma te dane wpisane w arkuszu od dawna. Czyli
handlowiec w terenie dostawał podpowiedź uboższą, niż mogła być.

DLACZEGO OSOBNE NARZĘDZIE, A NIE CZĘŚĆ IMPORTU
Import `PH Nowy` wciąga placówki i leady; ta zakładka to zupełnie inne dane
(ludzie, nie szkoły) i inna kadencja — rejony zmieniają się parę razy w roku,
a nie przy każdym imporcie. Do tego wymagają OGLĄDU: klient wpisuje je jako
swobodny tekst i połowa pozycji niesie komentarz („Knurów - nie odebrała
telefonu", „Chorzów (od grudnia powiat Mikołów)").

CZEGO NIE RUSZAMY
Kasia (08.08) wprost: z arkusza trenerów bierzemy tylko rejony oraz mail
i telefon — nie adres domowy, nie „ognisko", nie „nastawienie do pracy".
Telefonów i maili nie ma dziś gdzie zapisać (trener to pozycja słownika, nie
tabela z polami kontaktowymi), więc narzędzie ich nie dotyka; jeśli klient
będzie ich potrzebował w aplikacji, to osobna decyzja o schemacie.

UŻYCIE
    python narzedzia/trenerzy.py rejony --plik "PH PRÓBA Nowy...xlsx"
    python narzedzia/trenerzy.py rejony --plik "..." --zapisz --profil test
"""
import argparse
import os
import re
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

KATALOG = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, KATALOG)

PROFILE = ["prod", "test", "pusta"]
ZAKLADKA = "Trenerzy regiony"

# Dopiski, które klient wpisuje w tej samej komórce co rejon. To notatki
# robocze („nie odebrała telefonu"), nie nazwy miejscowości.
_SMIECI = re.compile(r"\?{2,}|nie odebra\w*(\s+telefonu)?|tylko duże gryp\w*"
                     r"|dwie pod rząd", re.IGNORECASE)


def _czysc_region(tekst):
    """Swobodny tekst klienta → lista kandydatów na miasta."""
    t = re.sub(r"\([^)]*\)", " ", tekst or "")     # komentarze w nawiasach
    t = _SMIECI.sub(" ", t)
    t = t.replace(";", "/").replace(",", "/")
    czesci = []
    for kawalek in t.split("/"):
        # „Knurów - nie odebrała telefonu" → „Knurów”; myślnik oddziela komentarz
        kawalek = re.split(r"\s+[-–]\s+", kawalek)[0]
        kawalek = kawalek.strip(" .?!")
        if kawalek:
            czesci.append(kawalek)
    return czesci


def dopasuj_miasto(kandydat, miasta, fold):
    """
    Kandydat → wartość ze słownika miast albo None.

    Klient pisze też nazwy szkół zamiast miast („SP 27 Katowice"), więc poza
    porównaniem wprost szukamy nazwy miasta WEWNĄTRZ tekstu — to ta sama
    informacja, tylko podana przez konkretną placówkę.
    """
    k = fold(kandydat)
    if not k:
        return None
    for m, czysta in miasta:
        if k == czysta:
            return m
    for m, czysta in miasta:
        if czysta and (czysta in k or k in czysta):
            return m
    return None


def zbierz(plik, profil):
    os.environ["PROFIL"] = profil
    import openpyxl                              # noqa: E402
    import db                                    # noqa: E402
    import parsers as P                          # noqa: E402

    conn = db.get_conn()
    slownik = db.slownik_values(conn, "trener")
    miasta = [(m, P._fold(P.strip_prefix(m)[1])) for m in db.slownik_values(conn, "miasto")]

    wb = openpyxl.load_workbook(plik, data_only=True)
    if ZAKLADKA not in wb.sheetnames:
        conn.close()
        raise SystemExit("Brak zakładki %r w pliku" % ZAKLADKA)
    ws = wb[ZAKLADKA]
    naglowki = [(c.value or "").strip().lower() if isinstance(c.value, str) else ""
                for c in ws[1]]
    kol = {h: i for i, h in enumerate(naglowki) if h}

    def pole(row, nazwa):
        i = kol.get(nazwa)
        v = row[i] if i is not None and i < len(row) else None
        return "" if v is None else str(v).strip()

    wynik, nierozpoznane, spoza_slownika = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nazwisko = pole(row, "nazwisko")
        region = pole(row, "region")
        if not nazwisko or not region:
            continue
        # trener ze słownika po nazwisku (słownik ma prefiksy „04. Zemela")
        nf = P._fold(nazwisko)
        trener = None
        for s in slownik:
            czysta = P._fold(P.strip_prefix(s)[1])
            if czysta == nf or nf in czysta.split():
                trener = s
                break
        if not trener:
            spoza_slownika.append((nazwisko, region))
            continue
        miasta_trenera, nieznane = [], []
        for kandydat in _czysc_region(region):
            m = dopasuj_miasto(kandydat, miasta, P._fold)
            if m and m not in miasta_trenera:
                miasta_trenera.append(m)
            elif not m:
                nieznane.append(kandydat)
        if nieznane:
            nierozpoznane.append((nazwisko, nieznane))
        if miasta_trenera:
            wynik.append((trener, miasta_trenera, region))
    wb.close()
    conn.close()
    return wynik, nierozpoznane, spoza_slownika


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    pod = ap.add_subparsers(dest="cmd", required=True)
    p = pod.add_parser("rejony", help="przenieś rejony z arkusza do tabeli")
    p.add_argument("--plik", required=True)
    p.add_argument("--profil", default="test", choices=PROFILE)
    p.add_argument("--zapisz", action="store_true",
                   help="bez tego tylko pokazuje, co by zrobił")
    a = ap.parse_args()
    if not os.path.exists(a.plik):
        raise SystemExit("Nie ma pliku: %s" % a.plik)

    wynik, nierozpoznane, spoza = zbierz(a.plik, a.profil)

    print("REJONY DO USTAWIENIA (%d trenerów)" % len(wynik))
    for trener, miasta, zrodlo in wynik:
        print("  %-30s → %s" % (trener[:30], ", ".join(miasta)))
        if len(miasta) != len(_czysc_region(zrodlo)):
            print("       (w arkuszu: %s)" % zrodlo[:70])

    if nierozpoznane:
        print("\nNIEROZPOZNANE FRAGMENTY (%d osób) — do sprawdzenia z Kasią:"
              % len(nierozpoznane))
        for nazwisko, co in nierozpoznane:
            print("  %-24s %s" % (nazwisko[:24], co))

    if spoza:
        print("\nOSOBY Z REJONEM, KTÓRYCH NIE MA W SŁOWNIKU TRENERÓW (%d):" % len(spoza))
        for nazwisko, region in spoza:
            print("  %-24s region: %s" % (nazwisko[:24], region[:40]))
        print("  → dodaj je w Słownikach (konto założy się samo) albo pomiń")

    if not a.zapisz:
        print("\nTo był podgląd. Zapis: dodaj --zapisz")
        return 0

    os.environ["PROFIL"] = a.profil
    import db                                    # noqa: E402
    import przydzial as pz                       # noqa: E402
    conn = db.get_conn()
    n = 0
    for trener, miasta, _ in wynik:
        n += pz.ustaw_rejon(conn, trener, miasta)
    db.zapisz_log(conn, kto="import rejonów", co="rejony trenerów",
                  po="%d trenerów, %d przypisań miast" % (len(wynik), n))
    conn.commit()
    conn.close()
    print("\nzapisano: %d trenerów, %d przypisań miast (profil %s)"
          % (len(wynik), n, a.profil))
    return 0


if __name__ == "__main__":
    sys.exit(main())
