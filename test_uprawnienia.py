# -*- coding: utf-8 -*-
"""
Testy własności rekordu przy ZAPISIE (P01, P02 — zgłoszenie K01 Kasi, 20.08.2026).

Zgłoszenie brzmiało: „Zablokuj PH możliwość edycji danych innego PH, bo teraz
mogą zmienić dosłownie wszystko". Przy sprawdzaniu okazało się, że jest gorzej,
niż wyglądało: wśród pól edytowalnych są `handlowiec` i `deadline`, więc
handlowiec mógł PRZYPISAĆ SOBIE cudzą szkołę i PRZEDŁUŻYĆ SOBIE termin, po
którym szkoła wraca do puli. W historii zmian wyglądało to jak zwykła praca.

Dlaczego to osobny plik testów: blokada w interfejsie już wcześniej „była" —
przycisków po prostu nie było widać. Zapis idzie zwykłym `fetch`, a numer leada
stoi w pasku adresu, więc jedyna blokada, która cokolwiek znaczy, siedzi przy
zapisie. Ten plik pilnuje właśnie tej warstwy i niczego innego.

Czego te testy NIE sprawdzają: podglądu. Kasia chce widzieć, kto miał szkołę
wcześniej i co z nią zrobił — odbieramy zapis, nie widok.

Uruchomienie:  python test_uprawnienia.py
Działa na WŁASNEJ, tymczasowej bazie.
"""
import json
import os
import sys
import tempfile

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

TMP = tempfile.mkdtemp(prefix="leady_v5_uprawnienia_test_")
os.environ["DATA_DIR"] = TMP

import app as A                      # noqa: E402
import db                            # noqa: E402
import uzytkownicy as uz             # noqa: E402
from seed import bootstrap           # noqa: E402

KL = A.app.test_client()

PH_A = "90. Test-Handlowiec-A"
PH_B = "91. Test-Handlowiec-B"
KOOR = "99. Test-Koordynator"
BIURO = "98. Test-Biuro"
PIN = {PH_A: "1111", PH_B: "2222", KOOR: "3333", BIURO: "4444"}

WYNIKI = []


def sprawdz(nazwa, warunek, opis=""):
    WYNIKI.append((nazwa, bool(warunek), opis))
    print("  [%s] %s%s" % ("OK  " if warunek else "BLAD", nazwa,
                           (" — " + opis) if opis else ""))
    return bool(warunek)


def zaloguj(osoba):
    """Przelogowanie w trakcie testu — każdy scenariusz zaczyna się od tego,
    KTO siedzi przy klawiaturze. To jest cała istota tych sprawdzeń."""
    r = KL.post("/api/logowanie", json={"osoba": osoba, "pin": PIN[osoba]})
    assert r.status_code == 200, "logowanie %s nie przeszło: %s" % (osoba, r.get_data())
    with KL.session_transaction() as s:
        s["csrf"] = "test-csrf"
    KL.environ_base["HTTP_X_CSRF"] = "test-csrf"


def patch(url, payload):
    r = KL.patch(url, data=json.dumps(payload), content_type="application/json")
    return r.status_code, (r.get_json() or {})


def post(url, payload):
    r = KL.post(url, data=json.dumps(payload), content_type="application/json")
    return r.status_code, (r.get_json() or {})


def usun(url):
    r = KL.delete(url)
    return r.status_code, (r.get_json() or {})


# --------------------------------------------------------------- przygotowanie

def przygotuj():
    conn = db.get_conn()
    bootstrap()
    conn = db.get_conn()
    uz.init(conn)
    for osoba in (PH_A, PH_B):
        conn.execute("INSERT INTO slowniki (rodzaj, wartosc, aktywny) VALUES ('handlowiec', ?, 1)",
                     (osoba,))
    conn.commit()
    for osoba, rola in ((PH_A, "handlowiec"), (PH_B, "handlowiec"),
                        (KOOR, "koordynator"), (BIURO, "biuro")):
        if not uz.znajdz(conn, osoba):
            uz.utworz(conn, osoba, rola, PIN[osoba])

    def szkola(nazwa, wlasciciel):
        cur = conn.execute("INSERT INTO placowki (nazwa, zrodlo) VALUES (?, 'test')", (nazwa,))
        pid = cur.lastrowid
        cur = conn.execute(
            "INSERT INTO leady (placowka_id, handlowiec, status_realizacji, deadline) "
            "VALUES (?,?,?,?)", (pid, wlasciciel, "01. Próba kontaktu (Brak konkretów)",
                                 "2026-09-30"))
        return cur.lastrowid

    ids = {
        "a": szkola("Szkoła A (moja)", PH_A),
        "b": szkola("Szkoła B (cudza)", PH_B),
        "niczyja": szkola("Szkoła bez opiekuna", None),
    }
    # po jednym wpisie w kalendarzu na szkołę A i B
    for klucz in ("a", "b"):
        cur = conn.execute(
            "INSERT INTO eventy (lead_id, typ, data, godz_od) VALUES (?, 'DT', '2026-09-15', '09:00')",
            (ids[klucz],))
        ids["event_" + klucz] = cur.lastrowid
    conn.commit()
    conn.close()
    return ids


# ------------------------------------------------------------------- scenariusze

def test_zapis_na_cudzym(ids):
    print("\n-- handlowiec przy CUDZEJ szkole --")
    zaloguj(PH_A)
    kod, odp = patch("/api/lead/%d" % ids["b"], {"field": "uwagi", "value": "podmieniam"})
    sprawdz("PATCH cudzego leada odrzucony", kod == 403, "kod %s" % kod)
    sprawdz("komunikat mówi, KTO prowadzi szkołę", PH_B in (odp.get("error") or ""),
            odp.get("error"))

    conn = db.get_conn()
    w_bazie = conn.execute("SELECT uwagi FROM leady WHERE id=?", (ids["b"],)).fetchone()["uwagi"]
    conn.close()
    sprawdz("cudzy rekord naprawdę nietknięty", not w_bazie, "uwagi=%r" % w_bazie)

    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "telefon", "value": "999888777"})
    sprawdz("pole PLACÓWKI na cudzej szkole też odrzucone", kod == 403, "kod %s" % kod)

    kod, _ = post("/api/pin", {"id": ids["b"], "pin": True})
    sprawdz("przypięcie cudzej szkoły na swój tydzień odrzucone", kod == 403, "kod %s" % kod)


def test_zapis_na_wlasnym(ids):
    print("\n-- handlowiec przy WŁASNEJ szkole (ma działać jak dotąd) --")
    zaloguj(PH_A)
    kod, _ = patch("/api/lead/%d" % ids["a"], {"field": "uwagi", "value": "dzwoniłem"})
    sprawdz("PATCH własnego leada przechodzi", kod == 200, "kod %s" % kod)

    conn = db.get_conn()
    w_bazie = conn.execute("SELECT uwagi FROM leady WHERE id=?", (ids["a"],)).fetchone()["uwagi"]
    conn.close()
    sprawdz("zapis naprawdę wszedł do bazy", w_bazie == "dzwoniłem", "uwagi=%r" % w_bazie)

    kod, _ = post("/api/pin", {"id": ids["a"], "pin": True})
    sprawdz("przypięcie własnej szkoły przechodzi", kod == 200, "kod %s" % kod)


def test_pola_zastrzezone(ids):
    print("\n-- pola, których handlowiec nie rusza nawet u siebie --")
    zaloguj(PH_A)
    kod, odp = patch("/api/lead/%d" % ids["a"], {"field": "handlowiec", "value": PH_A})
    sprawdz("przypisanie szkoły sobie samemu odrzucone", kod == 403, "kod %s" % kod)

    kod, _ = patch("/api/lead/%d" % ids["a"], {"field": "deadline", "value": "2027-12-31"})
    sprawdz("przedłużenie sobie terminu odrzucone", kod == 403, "kod %s" % kod)

    conn = db.get_conn()
    r = conn.execute("SELECT handlowiec, deadline FROM leady WHERE id=?", (ids["a"],)).fetchone()
    conn.close()
    sprawdz("termin w bazie bez zmian", r["deadline"] == "2026-09-30", r["deadline"])

    # to samo, ale na cudzej szkole — czyli próba przejęcia
    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "handlowiec", "value": PH_A})
    sprawdz("przejęcie cudzej szkoły odrzucone", kod == 403, "kod %s" % kod)
    conn = db.get_conn()
    r = conn.execute("SELECT handlowiec FROM leady WHERE id=?", (ids["b"],)).fetchone()
    conn.close()
    sprawdz("cudza szkoła dalej ma swojego opiekuna", r["handlowiec"] == PH_B, r["handlowiec"])


def test_szkola_niczyja(ids):
    print("\n-- szkoła bez opiekuna (przydziela koordynator, nie handlowiec) --")
    zaloguj(PH_A)
    kod, odp = patch("/api/lead/%d" % ids["niczyja"], {"field": "uwagi", "value": "biorę"})
    sprawdz("zapis na nieprzydzielonej szkole odrzucony", kod == 403, "kod %s" % kod)
    sprawdz("komunikat kieruje do koordynatora",
            "koordynator" in (odp.get("error") or "").lower(), odp.get("error"))


def test_kalendarz(ids):
    print("\n-- wpisy w kalendarzu należą do szkoły, nie do klikającego --")
    zaloguj(PH_A)
    kod, _ = usun("/api/event/%d" % ids["event_b"])
    sprawdz("skasowanie cudzego DT odrzucone", kod == 403, "kod %s" % kod)

    conn = db.get_conn()
    ile = conn.execute("SELECT COUNT(*) c FROM eventy WHERE id=?",
                       (ids["event_b"],)).fetchone()["c"]
    conn.close()
    sprawdz("cudze DT nadal jest w kalendarzu", ile == 1)

    kod, _ = patch("/api/event/%d" % ids["event_b"], {"field": "godz_od", "value": "07:00"})
    sprawdz("przestawienie godziny cudzego DT odrzucone", kod == 403, "kod %s" % kod)

    kod, _ = post("/api/event", {"lead_id": ids["b"], "typ": "DT", "data": "2026-10-01"})
    sprawdz("dopisanie DT do cudzej szkoły odrzucone", kod == 403, "kod %s" % kod)

    kod, _ = patch("/api/event/%d" % ids["event_a"], {"field": "godz_od", "value": "08:30"})
    sprawdz("własne DT dalej da się przestawić", kod == 200, "kod %s" % kod)


def test_odwolanie(ids):
    """P08 — decyzja z 20.08: handlowiec ODWOŁUJE, koordynator odwołuje I KASUJE.

    Podział nie jest kosmetyczny. Odwołanie zostawia powód, osobę i datę, więc
    zostaje ślad, że temat był i dlaczego się nie udał — to jest liczba, której
    Kasia szuka w raporcie wykonania. Kasowanie zabiera ten dowód, więc siedzi
    przy jednej osobie.
    """
    print("\n-- odwołanie kontra kasowanie --")
    conn = db.get_conn()
    ev_a = conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od, trener) "
                        "VALUES (?, 'DT', '2026-10-06', '09:00', '13. Cebula')",
                        (ids["a"],)).lastrowid
    ev_b = conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od) "
                        "VALUES (?, 'DT', '2026-10-07', '09:00')", (ids["b"],)).lastrowid
    conn.commit()
    conn.close()

    zaloguj(PH_A)
    kod, odp = post("/api/event/%d/odwolaj" % ev_a, {"powod": ""})
    sprawdz("odwołanie bez powodu odrzucone", kod == 400, odp.get("error"))

    kod, _ = post("/api/event/%d/odwolaj" % ev_b, {"powod": "szkoła odmówiła"})
    sprawdz("handlowiec nie odwoła cudzego spotkania", kod == 403, "kod %s" % kod)

    kod, odp = post("/api/event/%d/odwolaj" % ev_a,
                    {"powod": "dyrektor odwołał dzień przed"})
    sprawdz("handlowiec odwołuje WŁASNE spotkanie", kod == 200, "kod %s" % kod)

    conn = db.get_conn()
    r = conn.execute("SELECT odwolane, powod_odwolania, odwolal FROM eventy WHERE id=?",
                     (ev_a,)).fetchone()
    conn.close()
    sprawdz("wpis został w bazie, tylko oznaczony", bool(r and r["odwolane"]))
    sprawdz("zapisał się powód", "dyrektor odwołał" in (r["powod_odwolania"] or ""))
    sprawdz("zapisało się KTO odwołał", r["odwolal"] == PH_A, str(r["odwolal"]))

    kod, _ = post("/api/event/%d/odwolaj" % ev_a, {"powod": "jeszcze raz"})
    sprawdz("drugie odwołanie tego samego wpisu odrzucone", kod == 400, "kod %s" % kod)

    # Kasowanie bez śladu — handlowiec nie może, także u siebie.
    kod, _ = usun("/api/event/%d" % ev_a)
    sprawdz("handlowiec NIE skasuje nawet własnego spotkania", kod == 403, "kod %s" % kod)
    conn = db.get_conn()
    ile = conn.execute("SELECT COUNT(*) c FROM eventy WHERE id=?", (ev_a,)).fetchone()["c"]
    conn.close()
    sprawdz("wpis nadal jest w bazie", ile == 1)

    zaloguj(KOOR)
    kod, _ = post("/api/event/%d/odwolaj" % ev_a, {"cofnij": True})
    sprawdz("koordynator cofa odwołanie", kod == 200, "kod %s" % kod)
    conn = db.get_conn()
    r = conn.execute("SELECT odwolane FROM eventy WHERE id=?", (ev_a,)).fetchone()
    conn.close()
    sprawdz("po cofnięciu wpis znów jest aktualny", not r["odwolane"])

    kod, _ = usun("/api/event/%d" % ev_a)
    sprawdz("koordynator kasuje bez śladu", kod == 200, "kod %s" % kod)


def test_tworzenie_leada(ids):
    """
    ZAKŁADANIE PLACÓWEK PRZESZŁO DO KOORDYNATORA (Kasia, 24.08).

    Zgłoszenie: „usuń tę możliwość, bo to powoduje, że PH wpisują coś z ręki
    sami i będą się dublować rzeczy, a wpisują nazwy jak popadnie".

    W rundzie z 20.08 zamknęliśmy tu węższą dziurę — endpoint zostawał otwarty
    dla handlowca, tylko właściciel szedł z sesji zamiast z żądania. Teraz
    zamyka się cały endpoint, bo po przejściu bazy na rejestr RSPO „nie ma jej
    na liście" znaczy prawie zawsze „szukam nie w tym powiecie".

    Sprawdzamy OBIE drogi, bo w aplikacji są dwie: ekran „Baza" (`/api/lead`)
    i formularz terenowy (`/api/formularz` z blokiem `placowka`). Zamknięcie
    jednej bez drugiej nie znaczyłoby nic.
    """
    print("\n-- placówki zakłada KOORDYNATOR, nie handlowiec (Kasia, 24.08) --")
    zaloguj(PH_A)
    kod, odp = post("/api/lead", {"nazwa": "Podrzucona szkoła", "handlowiec": PH_B})
    sprawdz("handlowiec NIE założy placówki przez /api/lead", kod == 403, "kod %s" % kod)

    kod, odp = post("/api/formularz", {
        "placowka": {"nazwa": "Wpisana z ręki", "miejscowosc": "Psary"}})
    sprawdz("handlowiec NIE założy placówki przez formularz", kod == 403, "kod %s" % kod)
    sprawdz("komunikat kieruje do powiatu, zanim wyśle do koordynatorki",
            "powiat" in (odp.get("error") or "").lower(), odp.get("error"))

    conn = db.get_conn()
    ile = conn.execute("SELECT COUNT(*) c FROM placowki WHERE nazwa IN "
                       "('Podrzucona szkoła','Wpisana z ręki')").fetchone()["c"]
    conn.close()
    sprawdz("żaden z odrzuconych wierszy nie wszedł do bazy", ile == 0, "jest %d" % ile)

    # Formularz na ISTNIEJĄCEJ placówce ma działać jak dotąd — blokada dotyczy
    # zakładania, nie zapisu ustaleń. Inaczej zabralibyśmy handlowcowi pracę,
    # o którą w tym formularzu chodzi.
    kod, _ = post("/api/formularz", {"lead_id": ids["a"],
                                     "kontakt": {"telefon": "600100200"}})
    sprawdz("zapis ustaleń na własnej szkole przechodzi dalej", kod == 200, "kod %s" % kod)

    zaloguj(KOOR)
    kod, odp = post("/api/lead", {"nazwa": "Szkoła od koordynatorki"})
    sprawdz("koordynator zakłada placówkę bez zmian", kod == 200, "kod %s" % kod)


def test_biuro(ids):
    """
    Rola BIURO (Kasia, 31.08): „widok jak koordynator, ale tylko do odczytu".

    Blokada stoi na METODZIE HTTP, nie na liście endpointów. Lista wymagałaby
    dopisywania każdego nowego zapisu, a pierwszy pominięty byłby dziurą, o
    której nikt by nie wiedział — dokładnie jak „Usuń lead", który przez trzy
    tygodnie renderował się handlowcowi i kończył odmową dopiero po kliknięciu.

    Dlatego test NIE wylicza endpointów po jednym „na wszelki wypadek", tylko
    sprawdza REGUŁĘ: podgląd wszędzie tak, każda metoda inna niż GET nie.
    """
    print("\n-- BIURO: widzi wszystko, nie zapisuje nic --")
    zaloguj(BIURO)

    for sciezka in ("/pulpit", "/baza", "/leady", "/kalendarz", "/zbiorczy",
                    "/niewykorzystane", "/tydzien", "/slowniki", "/uzytkownicy"):
        sprawdz("BIURO widzi %s" % sciezka,
                KL.get(sciezka).status_code == 200,
                "kod %s" % KL.get(sciezka).status_code)
    sprawdz("BIURO widzi kartę cudzej szkoły",
            KL.get("/lead/%d" % ids["b"]).status_code == 200)
    # Eksport to GET i jest świadomie poza TYLKO_KOORDYNATOR — biuro ma
    # odpowiadać na pytania o grafik, więc plik do odpowiedzi też mu się należy.
    sprawdz("BIURO pobiera eksport XLSX", KL.get("/export.xlsx").status_code == 200)

    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "uwagi", "value": "z biura"})
    sprawdz("BIURO nie zapisze notatki", kod == 403, "kod %s" % kod)
    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "handlowiec", "value": PH_A})
    sprawdz("BIURO nie przypisze szkoły", kod == 403, "kod %s" % kod)
    r = KL.post("/api/event", data=json.dumps({"lead_id": ids["b"], "typ": "DT",
                                               "data": "2026-11-05"}),
                content_type="application/json")
    sprawdz("BIURO nie doda spotkania", r.status_code == 403, "kod %s" % r.status_code)
    kod, _ = usun("/api/lead/%d" % ids["b"])
    sprawdz("BIURO nie skasuje leada", kod == 403, "kod %s" % kod)
    r = KL.post("/api/przypisz", data=json.dumps({"ids": [ids["b"]],
                                                  "handlowiec": PH_A}),
                content_type="application/json")
    sprawdz("BIURO nie użyje przydziału", r.status_code == 403, "kod %s" % r.status_code)

    # Import zostaje SAMEMU koordynatorowi — nawet do podglądu. Samo wejście nic
    # nie zapisuje, ale PO TO się na ten ekran wchodzi; pokazany roli bez prawa
    # zapisu kończyłby się odmową dopiero po wybraniu pliku.
    sprawdz("BIURO nie wchodzi nawet na ekran importu",
            KL.get("/import").status_code in (302, 403),
            "kod %s" % KL.get("/import").status_code)

    # Sedno: reguła ma trzymać także tam, gdzie nikt nie wymienił endpointu.
    # `api_zwrot_podglad` to POST, który NICZEGO nie zmienia — a mimo to ma być
    # zamknięty, bo blokada idzie po metodzie, nie po tym, co endpoint robi.
    r = KL.post("/api/zwrot/podglad", data="{}", content_type="application/json")
    sprawdz("reguła obejmuje też zapisy, których nikt nie wymienił",
            r.status_code == 403, "kod %s" % r.status_code)

    # Konto BIURO ma WIDZIEĆ swoje ograniczenie. Inaczej człowiek klika,
    # dostaje odmowę i zgłasza błąd — a to nie błąd, tylko jego uprawnienia.
    html = KL.get("/pulpit").get_data(as_text=True)
    sprawdz("pasek u góry mówi wprost „tylko podgląd”", "tylko podgląd" in html)

    sprawdz("rola jest w słowniku ról (da się ją nadać na ekranie Konta)",
            "biuro" in uz.ROLE)
    sprawdz("i jest oznaczona jako tylko do odczytu",
            "biuro" in uz.ROLE_TYLKO_ODCZYT)


def test_koordynator(ids):
    print("\n-- koordynator: bez zmian, wolno mu wszystko --")
    zaloguj(KOOR)
    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "uwagi", "value": "przejrzane"})
    sprawdz("koordynator zapisuje na cudzej szkole", kod == 200, "kod %s" % kod)

    kod, _ = patch("/api/lead/%d" % ids["b"], {"field": "deadline", "value": "2026-10-15"})
    sprawdz("koordynator ustawia termin", kod == 200, "kod %s" % kod)

    kod, _ = patch("/api/lead/%d" % ids["niczyja"], {"field": "handlowiec", "value": PH_A})
    sprawdz("koordynator przypisuje szkołę", kod == 200, "kod %s" % kod)

    kod, _ = usun("/api/event/%d" % ids["event_b"])
    sprawdz("koordynator kasuje wpis z kalendarza", kod == 200, "kod %s" % kod)


def test_podglad_zostaje(ids):
    print("\n-- podgląd cudzych rekordów ZOSTAJE (Kasia chce widzieć historię) --")
    zaloguj(PH_A)
    r = KL.get("/lead/%d" % ids["b"])
    sprawdz("handlowiec otwiera kartę cudzej szkoły", r.status_code == 200,
            "kod %s" % r.status_code)
    r = KL.get("/api/placowki")
    sprawdz("lista placówek dalej dostępna", r.status_code == 200, "kod %s" % r.status_code)


def _xlsx_teksty(dane):
    """Wszystkie komórki skoroszytu jako jeden tekst — do prostych asercji."""
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(dane), read_only=True)
    czesci = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            czesci += [str(c) for c in row if c is not None]
    return " | ".join(czesci)


def test_eksport(ids):
    """Eksport dla PH (30.08): raport z filtrów TAK, ale przybity do własnych
    szkół — nazwisko z sesji, nie z adresu, więc podmiana parametru w URL
    niczego nie otwiera."""
    print("\n-- eksport XLSX: PH dostaje raport, ale tylko własnych szkół --")
    import urllib.parse
    zaloguj(PH_A)
    r = KL.get("/export.xlsx")
    sprawdz("handlowiec pobiera eksport (do 30.08 dostawał 403)",
            r.status_code == 200, "kod %s" % r.status_code)
    t = _xlsx_teksty(r.data)
    sprawdz("w pliku są jego szkoły", "Szkoła A (moja)" in t)
    sprawdz("cudzych szkół w pliku nie ma", "Szkoła B (cudza)" not in t)

    r = KL.get("/export.xlsx?handlowiec=" + urllib.parse.quote(PH_B))
    t = _xlsx_teksty(r.data)
    sprawdz("parametr handlowiec w URL nie otwiera cudzych leadów",
            "Szkoła B (cudza)" not in t and "Szkoła A (moja)" in t)

    zaloguj(KOOR)
    r = KL.get("/export.xlsx?handlowiec=" + urllib.parse.quote(PH_B))
    t = _xlsx_teksty(r.data)
    sprawdz("koordynator eksportuje wg dowolnego filtra", "Szkoła B (cudza)" in t)

    # Sprzeczność zakresu (pkt 17 z 30.08): konkretny handlowiec + zakładka
    # „Do rozdania" dawały ZAWSZE 0 wierszy. Zakres ma ustąpić filtrowi osoby.
    r = KL.get("/export.xlsx?handlowiec=%s&zakres=nieprzydzielone"
               % urllib.parse.quote(PH_A))
    t = _xlsx_teksty(r.data)
    sprawdz("handlowiec + „nieprzydzielone” nie daje już pustki",
            "Szkoła A (moja)" in t)
    r = KL.get("/baza?handlowiec=%s&zakres=nieprzydzielone"
               % urllib.parse.quote(PH_A))
    sprawdz("ekran /baza z tym samym filtrem pokazuje szkoły",
            "Szkoła A (moja)" in r.get_data(as_text=True))


def test_oddanie(ids):
    """Oddanie leada przez handlowca (pkt 12, 30.08): tylko własny, tylko
    z powodem; u koordynatora czerwona plakietka, przydzielenie ją gasi."""
    print("\n-- oddanie leada z powodem --")
    conn = db.get_conn()
    pid = conn.execute("INSERT INTO placowki (nazwa, zrodlo) VALUES ('Szkoła oddawana', 'test')").lastrowid
    lid = conn.execute(
        "INSERT INTO leady (placowka_id, handlowiec, status_realizacji) "
        "VALUES (?,?, '01. Próba kontaktu (Brak konkretów)')", (pid, PH_A)).lastrowid
    conn.commit(); conn.close()

    zaloguj(PH_B)
    kod, _ = post("/api/lead/%d/oddaj" % lid, {"powod": "to nie moja szkoła"})
    sprawdz("cudzej szkoły nie da się oddać", kod == 403, "kod %s" % kod)

    zaloguj(PH_A)
    kod, odp = post("/api/lead/%d/oddaj" % lid, {"powod": "   "})
    sprawdz("oddanie bez powodu odrzucone", kod == 400, "kod %s" % kod)
    kod, _ = post("/api/lead/%d/oddaj" % lid, {"powod": "dostałem przez przypadek"})
    sprawdz("oddanie z powodem przechodzi", kod == 200, "kod %s" % kod)

    conn = db.get_conn()
    l = dict(conn.execute("SELECT * FROM leady WHERE id=?", (lid,)).fetchone())
    conn.close()
    sprawdz("lead wrócił do puli (bez handlowca)", l["handlowiec"] is None)
    sprawdz("powód i osoba zapisane na czerwoną plakietkę",
            l["zwrot_powod"] == "dostałem przez przypadek" and l["zwrot_kto"] == PH_A
            and bool(l["zwrot_zgloszony"]))

    zaloguj(KOOR)
    r = KL.get("/baza?zakres=nieprzydzielone")
    html = r.get_data(as_text=True)
    sprawdz("koordynator widzi zgłoszenie na /baza",
            "Szkoła oddawana" in html and "oddana:" in html)

    kod, _ = post("/api/przypisz", {"ids": [lid], "handlowiec": PH_B})
    conn = db.get_conn()
    l = dict(conn.execute("SELECT * FROM leady WHERE id=?", (lid,)).fetchone())
    conn.close()
    sprawdz("przydzielenie gasi plakietkę",
            kod == 200 and l["zwrot_zgloszony"] is None and l["zwrot_powod"] is None)

    zaloguj(PH_A)
    kod, _ = post("/api/lead/%d/zwrot-rozpatrzony" % lid, {})
    sprawdz("„rozpatrzone” może kliknąć tylko koordynator", kod == 403, "kod %s" % kod)


def test_odwolanie_terminu_cyklu(ids):
    """Odwołanie JEDNYCH zajęć cyklu (pkt 21, 30.08) idzie przez tę samą
    kontrolę własności co reszta zapisów — nowy endpoint nie może być obejściem."""
    print("\n-- odwołanie pojedynczego terminu cyklu: czyja szkoła --")
    conn = db.get_conn()
    cykle = {}
    for klucz in ("a", "b"):
        cykle[klucz] = conn.execute(
            "INSERT INTO eventy (lead_id, typ, data, godz_od, co_ile_tygodni) "
            "VALUES (?, 'CYKLICZNE', '2026-09-15', '12:00', 1)", (ids[klucz],)).lastrowid
    conn.commit(); conn.close()

    zaloguj(PH_A)
    kod, _ = post("/api/event/%d/odwolaj-termin" % cykle["b"],
                  {"data": "2026-09-22", "powod": "próba na cudzej"})
    sprawdz("cudzego terminu handlowiec nie odwoła", kod == 403, "kod %s" % kod)
    kod, _ = post("/api/event/%d/odwolaj-termin" % cykle["a"],
                  {"data": "2026-09-22", "powod": "wywiadówka"})
    sprawdz("swój termin odwołuje", kod == 200, "kod %s" % kod)
    kod, _ = post("/api/event/%d/odwolaj-termin" % cykle["a"], {"powod": "bez daty"})
    sprawdz("bez daty terminu odmowa", kod == 400, "kod %s" % kod)

    conn = db.get_conn()
    w = conn.execute("SELECT odwolal FROM wyjatki_cyklu WHERE event_id=?",
                     (cykle["a"],)).fetchone()
    conn.close()
    sprawdz("podpis bierze się z sesji, nie z żądania", w and w["odwolal"] == PH_A,
            str(dict(w) if w else None))

    # Przesuwanie terminów (pkt 18) idzie tą samą kontrolą — handlowiec MUSI
    # móc na swojej szkole, bo brak tej możliwości był powodem zgłoszenia
    # („PH wpisał cykle jako DT, bo nie mógł edytować dat").
    print("\n-- przesuwanie terminów cyklu: czyja szkoła --")
    kod, _ = post("/api/event/%d/termin" % cykle["b"],
                  {"data": "2026-09-15", "data_nowa": "2026-09-16"})
    sprawdz("cudzego terminu handlowiec nie przesunie", kod == 403, "kod %s" % kod)
    kod, _ = post("/api/event/%d/termin" % cykle["a"],
                  {"data": "2026-09-15", "data_nowa": "2026-09-16"})
    sprawdz("swój termin przesuwa", kod == 200, "kod %s" % kod)
    zaloguj(KOOR)
    kod, _ = post("/api/event/%d/termin" % cykle["b"],
                  {"data": "2026-09-15", "data_nowa": "2026-09-17"})
    sprawdz("koordynator przesuwa każdy", kod == 200, "kod %s" % kod)


def main():
    print("=" * 62)
    print("UPRAWNIENIA — właściciel rekordu przy zapisie (P01, P02)")
    print("=" * 62)
    ids = przygotuj()
    test_zapis_na_cudzym(ids)
    test_zapis_na_wlasnym(ids)
    test_pola_zastrzezone(ids)
    test_szkola_niczyja(ids)
    test_kalendarz(ids)
    test_odwolanie(ids)
    test_tworzenie_leada(ids)
    test_koordynator(ids)
    test_podglad_zostaje(ids)
    test_eksport(ids)
    test_oddanie(ids)
    test_odwolanie_terminu_cyklu(ids)
    test_biuro(ids)

    ok = sum(1 for _, w, _ in WYNIKI if w)
    print("\n" + "=" * 62)
    print("WYNIK: %d/%d sprawdzeń OK" % (ok, len(WYNIKI)))
    print("=" * 62)
    return 0 if ok == len(WYNIKI) else 1


if __name__ == "__main__":
    sys.exit(main())
