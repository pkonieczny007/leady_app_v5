# -*- coding: utf-8 -*-
"""
Nadanie numerów RSPO placówkom, które ich nie mają — etap M3 projektu
`docs/poprawka 23.08.2026/PROJEKT_BAZY_RSPO.md`.

PO CO NUMER, SKORO POWIAT JUŻ DZIAŁA
Powiat dało się wywieść z nazwy miejscowości (etap M5) i to wystarczyło do
filtrów. Numer RSPO jest potrzebny do trzech rzeczy, których obejść się nie da:

  1. Dołożenie brakujących SZKÓŁ. W rejestrze jest ich w naszych obszarach 549,
     mamy 539 — ale bez numeru nie sposób powiedzieć, które z tych 549 już
     mamy. Dołożenie „na oko" zrobiłoby ~540 dubli dokładnie tam, gdzie
     handlowcy mają całą swoją pracę.
  2. Prawdziwa miejscowość dla 68 rekordów, które siedziały w workach
     `09. Pszczyna` i `15. Będzin`. Ich miejscowość została po migracji
     przybliżona nazwą miasta powiatowego, bo w danych klienta nigdy jej nie
     było — urwała się przy imporcie razem ze słowem „powiat".
  3. Comiesięczne odświeżenie rejestru. Bez numeru nie ma po czym dopasować
     zmiany nazwy czy telefonu.

CZEGO AUTOMAT NIE ZROBI
Nie wpisze numeru tam, gdzie dwa źródła się nie zgadzają. Złe powiązanie nie
boli od razu — boli przez lata, bo comiesięczne odświeżanie będzie „poprawiać"
nie tę szkołę, a nikt nie skojarzy przyczyny. Rozjazd zawsze idzie do człowieka.

DWA ŹRÓDŁA
  · nasze dopasowanie po nazwie i geografii (kod niżej),
  · plik klienta `POPRAWKA BAZY - RSPO dopasowane.xlsx`, kluczowany po NASZYM
    `id` — więc nie wymaga zgadywania, tylko sprawdzenia, czy nazwa się zgadza
    (plik powstał 20.08 na kopii produkcji; gdyby id się przesunęły, wpisanie
    numeru po samym id podpięłoby szkołę pod cudzy rekord).
"""
import collections
import re

_RE_NUMER = re.compile(r"(?:^|\s|\b)(?:M|Z)?(?:SP|PM|PP|PS|ZSP|ZPO|ZS)\s*(?:NR\s*)?(\d+)",
                       re.IGNORECASE)
_RE_NUMER_REJESTR = re.compile(r"\bNR\s+(\d+)", re.IGNORECASE)

# Werdykty. Kolejność od najmocniejszego — tak też sortuje się raport.
WPISYWANE = ("zgodne", "pewne")


def _fold(s):
    s = (s or "").lower().translate(str.maketrans("ąćęłńóśźż", "acelnoszz"))
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _numer(nazwa, z_rejestru=False):
    m = (_RE_NUMER_REJESTR if z_rejestru else _RE_NUMER).search(nazwa or "")
    return int(m.group(1)) if m else None


def kategoria(nazwa, typ=""):
    """
    Zgrubny podział szkoła/przedszkole/inne — ale bez niego „MSP 2" w Knurowie
    dopasowywało się do BRANŻOWEJ SZKOŁY I STOPNIA NR 2 w tym samym mieście.
    Trzy kubełki wystarczą, bo dopasowujemy w obrębie jednej miejscowości.
    """
    t, n = _fold(typ), _fold(nazwa)
    if "przedszkol" in t or "przedszkol" in n or re.match(r"^(pm|pp|ps) ?\d", n):
        return "przedszkole"
    if "szkola podstawowa" in t or "szkola podstawowa" in n \
            or re.match(r"^m?sp ?\d", n) or re.match(r"^(zsp|zpo) ?\d", n):
        return "podstawowa"
    return "inne"


def _indeksy(conn):
    """Cztery sposoby trafienia w rekord rejestru — od najpewniejszego."""
    po_nazwie_miejsc = collections.defaultdict(list)
    po_nazwie_powiat = collections.defaultdict(list)
    po_numerze = collections.defaultdict(list)
    po_numerze_miejsc = collections.defaultdict(list)
    for r in conn.execute("SELECT rspo, nazwa, typ, powiat, miejscowosc "
                          "FROM rspo_rejestr WHERE nieobecna_od IS NULL"):
        kat = kategoria(r["nazwa"], r["typ"])
        po_nazwie_miejsc[(_fold(r["nazwa"]), _fold(r["miejscowosc"]))].append(r)
        po_nazwie_powiat[(_fold(r["nazwa"]), _fold(r["powiat"]))].append(r)
        nr = _numer(r["nazwa"], z_rejestru=True)
        if nr is not None:
            po_numerze[(_fold(r["powiat"]), nr, kat)].append(r)
            po_numerze_miejsc[(_fold(r["miejscowosc"]), nr, kat)].append(r)
    return po_nazwie_miejsc, po_nazwie_powiat, po_numerze, po_numerze_miejsc


def _po_nazwie_wlasnej(conn, nazwa, miejscowosc):
    """
    Rekordy rejestru z tej samej miejscowości, których nazwa zawiera WSZYSTKIE
    znaczące słowa naszej. Ta sama reguła, którą `dokladanie` odmawia dołożenia
    dubla — trzymamy ją w jednym miejscu, żeby nie rozjechały się dwie kopie.
    """
    import dokladanie
    slowa = dokladanie._slowa_znaczace(nazwa, miejscowosc)
    if not slowa:
        return []
    out = []
    for r in conn.execute(
            "SELECT rspo, nazwa, typ, powiat, miejscowosc FROM rspo_rejestr "
            "WHERE miejscowosc = ? AND nieobecna_od IS NULL", (miejscowosc,)):
        if slowa <= set(dokladanie._fold(r["nazwa"]).split()):
            out.append(r)
    return out


def _po_wsi_w_nazwie(conn, nazwa, powiat, kat):
    """
    Kandydaci z miejscowości, której nazwa siedzi w NAZWIE naszego rekordu.

    Bierze się to z worków powiatowych: „Sp Góra" i „Sp Miedźna" mają
    miejscowość `Pszczyna` (bo tyle zostało po urwanym słowie „powiat"),
    ale handlowiec dopisał wieś do nazwy szkoły. Szukamy więc nazwy wsi
    w tekście — od najdłuższej, żeby „Wisła Wielka" wygrała z „Wisłą".
    """
    igla = _fold(nazwa)
    wsie = sorted({r[0] for r in conn.execute(
        "SELECT DISTINCT miejscowosc FROM rspo_rejestr WHERE powiat=? "
        "AND miejscowosc IS NOT NULL AND miejscowosc <> ''", (powiat,))},
        key=lambda s: -len(s))
    for wies in wsie:
        f = _fold(wies)
        # `w mizerowie` nie zawiera `mizerow`, więc porównujemy po RDZENIU:
        # odmiana zmienia końcówkę, nie początek.
        rdzen = f[:max(4, len(f) - 2)]
        if len(rdzen) < 4 or rdzen not in igla:
            continue
        kand = [r for r in conn.execute(
            "SELECT rspo, nazwa, typ, miejscowosc FROM rspo_rejestr "
            "WHERE powiat=? AND lower(miejscowosc)=lower(?) AND nieobecna_od IS NULL",
            (powiat, wies))]
        kand = [k for k in kand if kategoria(k["nazwa"], k["typ"]) == kat]
        if kand:
            return kand
    return []


def _z_pliku_klienta(sciezka):
    """{id naszego rekordu: (numer rspo, nazwa z pliku)} — do kontroli krzyżowej."""
    if not sciezka:
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(sciezka, read_only=True, data_only=True)
    ws = wb["Szkoły"] if "Szkoły" in wb.sheetnames else wb[wb.sheetnames[0]]
    out, naglowki = {}, None
    for wiersz in ws.iter_rows(values_only=True):
        if naglowki is None:
            naglowki = [str(c or "").strip() for c in wiersz]
            continue
        d = dict(zip(naglowki, wiersz))
        try:
            ident = int(str(d.get("id") or "").strip())
            numer = int(str(d.get("Nr RSPO") or "").strip())
        except (TypeError, ValueError):
            continue
        out[ident] = (numer, str(d.get("Nazwa placówki") or "").strip())
    wb.close()
    return out


def dopasuj(conn, plik_klienta=None):
    """
    Lista wierszy decyzyjnych — po jednym na placówkę bez numeru RSPO.

    Kolejność prób jest kolejnością pewności:
      1. nazwa + miejscowość zgodne co do znaku,
      2. nazwa + POWIAT — dla 68 rekordów z worków powiatowych miejscowość jest
         przybliżeniem, więc krok 1 by ich nie złapał, a nazwa urzędowa
         w obrębie powiatu jest praktycznie unikalna,
      3. numer szkoły + powiat + kategoria, gdy w rejestrze jest DOKŁADNIE
         jeden taki kandydat.
    """
    po_nm, po_np, po_nr, po_nr_miejsc = _indeksy(conn)
    klient = _z_pliku_klienta(plik_klienta)

    wynik = []
    for p in conn.execute(
            "SELECT id, nazwa, miejscowosc, powiat, typ FROM placowki "
            "WHERE rspo IS NULL OR rspo = '' ORDER BY powiat, miejscowosc, nazwa"):
        nazwa = (p["nazwa"] or "").strip()
        w = {"id": p["id"], "nazwa": nazwa, "miejscowosc": p["miejscowosc"] or "",
             "powiat": p["powiat"] or "", "nasz": None, "jak": "", "nazwa_rspo": "",
             "klient": None, "alternatywy": "", "werdykt": "brak"}

        trafienia = po_nm.get((_fold(nazwa), _fold(p["miejscowosc"]))) or []
        jak = "nazwa+miejscowość"
        if len(trafienia) != 1:
            trafienia = po_np.get((_fold(nazwa), _fold(p["powiat"]))) or []
            jak = "nazwa+powiat"
        if len(trafienia) != 1:
            nr = _numer(nazwa)
            kat = kategoria(nazwa, p["typ"] or "")
            if nr is not None:
                # Najpierw po MIEJSCOWOŚCI — od czasu, gdy worki powiatowe
                # zostały rozpakowane, miejscowość znów coś znaczy i „Sp 1"
                # w Woli jest jednoznaczne, choć „SP nr 1" w powiecie
                # pszczyńskim jest kilka.
                trafienia = po_nr_miejsc.get((_fold(p["miejscowosc"]), nr, kat)) or []
                jak = "numer w nazwie+miejscowość"
                if len(trafienia) != 1:
                    trafienia = po_nr.get((_fold(p["powiat"]), nr, kat)) or []
                    jak = "numer w nazwie+powiat"
            else:
                trafienia = []

        if len(trafienia) != 1:
            # Nazwa własna placówki wewnątrz nazwy urzędowej. „Zając Poziomka"
            # to w rejestrze „NIEPUBLICZNE PRZEDSZKOLE ZAJĄC POZIOMKA
            # W DĄBROWIE GÓRNICZEJ". Tej samej reguły używamy przy dokładaniu,
            # żeby ODMÓWIĆ utworzenia dubla — skoro dowód jest dość mocny, by
            # wstrzymać zapis, jest dość mocny i na powiązanie.
            trafienia = _po_nazwie_wlasnej(conn, nazwa, p["miejscowosc"])
            jak = "nazwa własna w nazwie urzędowej"
        if len(trafienia) != 1 and not trafienia:
            # OSTATNIA PRÓBA: miejscowość schowana w NAZWIE, nie w kolumnie.
            # „Sp Miedźna", „Sp Góra", „Sp Frydyk" siedzą pod miejscowością
            # `Pszczyna`, bo tyle zostało z worka powiatowego — ale nazwę wsi
            # handlowiec dopisał do nazwy szkoły. Z tego NIGDY nie wpisujemy
            # numeru automatem: to podpowiedź do pliku decyzyjnego, żeby Kasia
            # wybierała z trzech kandydatów zamiast szukać od zera.
            trafienia = _po_wsi_w_nazwie(conn, nazwa, p["powiat"],
                                         kategoria(nazwa, p["typ"] or ""))
            if trafienia:
                w["alternatywy"] = " | ".join(
                    "%s → %s (%s)" % (t["rspo"], t["nazwa"], t["miejscowosc"])
                    for t in trafienia[:5])
                w["werdykt"] = "niepewne"
            trafienia = []              # podpowiedź, nie trafienie

        if len(trafienia) == 1:
            w["nasz"] = trafienia[0]["rspo"]
            w["nazwa_rspo"] = trafienia[0]["nazwa"]
            w["jak"] = jak
        elif trafienia:
            w["alternatywy"] = " | ".join("%s → %s" % (t["rspo"], t["nazwa"])
                                          for t in trafienia[:5])
            w["werdykt"] = "niepewne"

        z_pliku = klient.get(p["id"])
        if z_pliku:
            # Plik powstał na kopii produkcji z 20.08 i jest kluczowany po id.
            # Gdyby id się przesunęły, wpisanie numeru po samym id podpięłoby
            # szkołę pod cudzy rekord — dlatego wpierw sprawdzamy nazwę.
            if _fold(z_pliku[1]) == _fold(nazwa):
                w["klient"] = z_pliku[0]
            else:
                w["alternatywy"] = (w["alternatywy"] + " ; " if w["alternatywy"] else "") \
                    + "plik klienta ma pod tym id inną placówkę: %s" % z_pliku[1]

        if w["nasz"] and w["klient"]:
            w["werdykt"] = "zgodne" if w["nasz"] == w["klient"] else "rozjazd"
        elif w["nasz"]:
            w["werdykt"] = "pewne"
        elif w["klient"]:
            # Sam plik klienta to za mało: jego dopasowania nie sprawdzaliśmy,
            # a numer wpisany omyłkowo zwiąże nas z cudzą szkołą na lata.
            w["werdykt"] = "tylko plik klienta"
        wynik.append(w)

    # DUBLE WEWNĄTRZ NASZEJ BAZY — dopiero numer daje wspólny punkt odniesienia.
    # Dwa nasze rekordy wskazujące ten sam numer to ta sama szkoła zapisana
    # dwa razy: raz pełną nazwą, raz skrótem handlowca. Numeru nie dostaje ŻADEN
    # z nich — na jednym wisi historia, na drugim jedyne umówione DT, więc wybór
    # nie jest techniczny i należy do człowieka (etap M4).
    licznik = collections.Counter(w["nasz"] for w in wynik if w["nasz"])
    zajete = {int(r[0]) for r in conn.execute(
        "SELECT rspo FROM placowki WHERE rspo IS NOT NULL AND rspo <> ''")}

    # ROZSTRZYGNIĘCIE PARY: numer dostaje rekord o nazwie ZGODNEJ Z REJESTREM.
    # To ta sama reguła, którą i tak przewiduje scalanie dubli (M4): zostaje
    # rekord z pełną nazwą, skrót handlowca („MSP 1", „ZS 3 Rybnik") idzie do
    # scalenia. Bez tego para blokuje się nawzajem i placówka, którą MAMY,
    # wygląda w rozliczeniu z rejestrem na brakującą — a to jedyne 6 wierszy,
    # o które nie zgadzały się liczby.
    #
    # Rozstrzygamy TYLKO wtedy, gdy zgodność nazwy ma dokładnie jeden rekord.
    # Dwa jednakowo pasujące to prawdziwy remis i idą do człowieka.
    zgodni = collections.defaultdict(list)
    for w in wynik:
        if w["nasz"] and licznik[w["nasz"]] > 1 \
                and _fold(w["nazwa"]) == _fold(w["nazwa_rspo"]):
            zgodni[w["nasz"]].append(w["id"])
    rozstrzygniete = {numer: ident[0] for numer, ident in zgodni.items()
                      if len(ident) == 1}

    for w in wynik:
        if w["nasz"] and licznik[w["nasz"]] > 1:
            if rozstrzygniete.get(w["nasz"]) == w["id"]:
                w["jak"] = (w["jak"] or "") + " (nazwa zgodna z rejestrem)"
                w["werdykt"] = "zgodne" if w["klient"] == w["nasz"] else "pewne"
                continue
            w["werdykt"] = "dubel"
            w["alternatywy"] = ("%d nasze rekordy wskazują ten sam numer%s"
                                % (licznik[w["nasz"]],
                                   "; numer dostał id %d (nazwa zgodna z rejestrem), "
                                   "ten wiersz do scalenia"
                                   % rozstrzygniete[w["nasz"]]
                                   if w["nasz"] in rozstrzygniete else ""))
            w["nasz"] = None if w["nasz"] in rozstrzygniete else w["nasz"]
        elif w["nasz"] and w["nasz"] in zajete:
            # Numer wzięty już przez inną placówkę (np. dołożoną z rejestru).
            w["werdykt"] = "numer zajęty"
    return wynik


def wpisz(conn, wiersze, kto="migracja-rspo"):
    """
    Wpisuje numery tylko tam, gdzie werdykt na to pozwala. Każdy UPDATE zostawia
    ślad w `log`, żeby dało się to odkręcić listą, a nie z pamięci.
    """
    import db
    n = 0
    for w in wiersze:
        if w["werdykt"] not in WPISYWANE or not w["nasz"]:
            continue
        conn.execute("UPDATE placowki SET rspo=?, updated_at=datetime('now') WHERE id=?",
                     (str(w["nasz"]), w["id"]))
        lead = conn.execute("SELECT id FROM leady WHERE placowka_id=? ORDER BY id LIMIT 1",
                            (w["id"],)).fetchone()
        db.zapisz_log(conn, lead_id=lead["id"] if lead else None, kto=kto,
                      co="migracja-rspo: numer", pole="rspo", przed=None,
                      po="%s (%s)" % (w["nasz"], w["jak"]))
        n += 1
    conn.commit()
    return n


def _kontekst_rekordu(conn, ident):
    """
    Czym ten rekord ŻYJE — i czy w bazie stoi już jego bliźniak z rejestru.

    Bez tego plik decyzyjny odpowiada tylko na „jaki numer", a człowiek i tak
    musi wejść do aplikacji, żeby sprawdzić, czy wolno rekord scalić. Tu widać
    od razu: kto go prowadzi, ile ma umówionych spotkań i który rekord jest
    kandydatem na cel scalenia (ten z pełną nazwą i numerem RSPO).
    """
    d = {"handlowiec": "", "status": "", "eventy": 0, "adres": "", "telefon": "",
         "blizniak": ""}
    lead = conn.execute("""
        SELECT l.handlowiec, l.status_realizacji, p.adres, p.telefon, p.miejscowosc,
               p.nazwa, (SELECT COUNT(*) FROM eventy e WHERE e.lead_id = l.id) n
          FROM placowki p LEFT JOIN leady l ON l.placowka_id = p.id
         WHERE p.id = ? ORDER BY l.id LIMIT 1""", (ident,)).fetchone()
    if not lead:
        return d
    d.update(handlowiec=lead["handlowiec"] or "", status=lead["status_realizacji"] or "",
             eventy=lead["n"] or 0, adres=lead["adres"] or "", telefon=lead["telefon"] or "")

    # Bliźniak: rekord Z NUMEREM w tej samej miejscowości, którego nazwa zawiera
    # wszystkie znaczące słowa naszej (albo ma ten sam numer szkoły). To ta sama
    # reguła, którą dokładanie odmawia utworzenia dubla.
    import dokladanie
    slowa = dokladanie._slowa_znaczace(lead["nazwa"], lead["miejscowosc"])
    nr = dokladanie.numer_szkoly(lead["nazwa"])
    for r in conn.execute(
            "SELECT id, nazwa, rspo FROM placowki WHERE miejscowosc = ? "
            "AND rspo IS NOT NULL AND rspo <> '' AND id <> ?",
            (lead["miejscowosc"], ident)):
        slowa_r = set(dokladanie._fold(r["nazwa"]).split())
        nr_r = dokladanie.numer_szkoly(r["nazwa"])
        pasuje = (slowa and slowa <= slowa_r) or (nr and nr_r and nr == nr_r)
        if pasuje:
            d["blizniak"] = "id %s · RSPO %s · %s" % (r["id"], r["rspo"], r["nazwa"])
            break
    return d


def do_xlsx(conn, wiersze, docelowy):
    """Plik decyzyjny dla koordynatorki — tylko to, czego automat nie ruszył."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Do sprawdzenia"
    naglowki = ["id", "Nazwa u nas", "Miejscowość", "Powiat", "Typ / adres",
                "Handlowiec", "Status", "Spotkań", "Werdykt automatu",
                "Nasz numer", "Numer z pliku klienta", "Nazwa w rejestrze",
                "Kandydaci / uwagi",
                "BLIŹNIAK w bazie (kandydat do scalenia)",
                "DECYZJA — numer RSPO", "DECYZJA — scalić z id"]
    ws.append(naglowki)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    zolty = PatternFill("solid", fgColor="FFF2CC")

    for w in wiersze:
        if w["werdykt"] in WPISYWANE:
            continue
        k = _kontekst_rekordu(conn, w["id"])
        ws.append([w["id"], w["nazwa"], w["miejscowosc"], w["powiat"],
                   (k["adres"] + (" · tel. " + k["telefon"] if k["telefon"] else "")).strip(),
                   k["handlowiec"], k["status"], k["eventy"], w["werdykt"],
                   w["nasz"] or "", w["klient"] or "", w["nazwa_rspo"],
                   w["alternatywy"], k["blizniak"], "", ""])
        # Wiersz z umówionym spotkaniem świeci: na nim wisi praca, więc przy
        # scalaniu to ON jest tym, z którego NIC nie może zginąć.
        if k["eventy"]:
            for c in ws[ws.max_row]:
                c.fill = zolty

    for kol, szer in zip(
            ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"],
            (6, 40, 16, 14, 34, 16, 26, 8, 16, 11, 13, 44, 52, 52, 18, 18)):
        ws.column_dimensions[kol].width = szer
    ws.freeze_panes = "B2"
    wb.save(docelowy)
    return docelowy


def wczytaj_decyzje(sciezka):
    """Plik po ręcznym uzupełnieniu → [{id, nasz, jak, werdykt}] gotowe dla `wpisz`."""
    import openpyxl
    wb = openpyxl.load_workbook(sciezka, read_only=True, data_only=True)
    ws = wb.active
    out, naglowki = [], None
    for wiersz in ws.iter_rows(values_only=True):
        if naglowki is None:
            naglowki = [str(c or "").strip() for c in wiersz]
            continue
        d = dict(zip(naglowki, wiersz))
        try:
            ident = int(str(d.get("id") or "").strip())
            numer = int(str(d.get("DECYZJA — wpisz numer RSPO") or "").strip())
        except (TypeError, ValueError):
            continue
        out.append({"id": ident, "nasz": numer, "jak": "decyzja człowieka",
                    "werdykt": "zgodne"})
    wb.close()
    return out
