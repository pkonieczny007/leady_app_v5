# -*- coding: utf-8 -*-
"""
Testy mostu do aplikacji partnerskiej (Akademia Silesia3D).

Trzy rzeczy pilnowane tu twardo, bo każda z nich jest błędem, którego NIKT nie
zauważy gołym okiem:

1. ZAPORA NA DANE OSOBOWE. Telefon, mail, osoba kontaktowa i notatki handlowca
   nie mają prawa wyjść. Szukamy po WARTOŚCIACH wpisanych do bazy testowej,
   a nie po nazwach pól — przemianowanie kolumny nie ma oszukać testu, a wolny
   tekst potrafi wynieść numer telefonu w polu, które nazywa się „uwagi".

2. ZGODNOŚĆ Z KALENDARZEM. Wystąpienia cyklu w pliku muszą być tymi samymi
   datami, które widzi koordynator na ekranie. Rozjazd między tym, co widzi
   klient, a tym, co dostaje partner, wychodzi dopiero przy awanturze
   o nieodbyte zajęcia.

3. ATOMOWOŚĆ. Odbiorca czyta plik w losowym momencie. Nigdy nie ma prawa
   zobaczyć połowy.
"""
import datetime as dt
import json
import os
import sys
import tempfile

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
os.environ["DATA_DIR"] = tempfile.mkdtemp()
os.environ["MOST_DIR"] = tempfile.mkdtemp()
os.environ["PROFIL"] = "test"

import calendar_view as cv     # noqa: E402
import db                      # noqa: E402
import most                    # noqa: E402
from most import dane, pliki   # noqa: E402

WYNIKI = []

# Dane osobowe wpisane do bazy PO TO, żeby sprawdzić, że nie wychodzą.
# Wartości są nietypowe celowo — przypadkowe trafienie w „Kowalski" nic by nie
# dowiodło, a te ciągi nie mają prawa pojawić się w pliku z żadnego innego powodu.
TELEFON = "601-000-777"
MAIL = "dyrekcja.tajna@przyklad-testowy.pl"
OSOBA = "Bogumiła Niewychodząca"
UWAGI_LEADA = "dzwonić po 15, prywatny numer 502-111-222"
UWAGI_EVENTU = "wejście od podwórza, kod do bramy 4821"


def sprawdz(nazwa, warunek, opis=""):
    WYNIKI.append((nazwa, bool(warunek)))
    print("  [%s] %s%s" % ("OK  " if warunek else "BLAD", nazwa,
                           (" — " + opis) if opis else ""))
    return bool(warunek)


def _slownik(conn, rodzaj, wartosci):
    for i, w in enumerate(wartosci):
        conn.execute("INSERT OR IGNORE INTO slowniki (rodzaj, wartosc, sort_order)"
                     " VALUES (?,?,?)", (rodzaj, w, i))
    conn.commit()


def zbuduj_baze():
    """
    Jedna placówka, jeden lead, cztery eventy — po jednym na każdy przypadek,
    który most musi umieć: cykl z reguły, cykl z listy dat, DT, cykl odwołany.
    """
    conn = db.get_conn()
    db.init_db(conn)
    _slownik(conn, "trener", ["05. Kowalski", "07. Nowak"])

    pid = conn.execute(
        "INSERT INTO placowki (nazwa, typ, miejscowosc, adres, osoba_kontakt, "
        "telefon, mail, powiat, gmina, rspo) VALUES (?,?,?,?,?,?,?,?,?,?)",
        ("Szkoła Podstawowa nr 7", "01. Szkoła podstawowa", "Zabrze", "ul. Testowa 1",
         OSOBA, TELEFON, MAIL, "Zabrze", "Zabrze", "15109")).lastrowid
    lid = conn.execute(
        "INSERT INTO leady (placowka_id, handlowiec, uwagi) VALUES (?,?,?)",
        (pid, "01. Sacawa", UWAGI_LEADA)).lastrowid

    # 1. cykl z REGUŁY — poniedziałek 05.10.2026, co tydzień
    cykl = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, godz_od, godz_do, trener, drukarz,"
        " numer_sali, grupa, ilosc_klas, ilosc_dzieci, co_ile_tygodni, uwagi)"
        " VALUES (?,'CYKLICZNE','2026-10-05','14:00','15:00',?,?,?,?,?,?,1,?)",
        (lid, "05. Kowalski", "07. Nowak", "12", "A", 3, 60, UWAGI_EVENTU)).lastrowid
    # jedno wystąpienie odwołane
    conn.execute(
        "INSERT INTO wyjatki_cyklu (event_id, data, odwolane, powod_odwolania, odwolal)"
        " VALUES (?,?,1,?,?)", (cykl, "2026-10-19", "ferie w szkole", "01. Sacawa"))

    # 2. cykl z LISTĄ DAT (pakiet przedszkolny) — reguła mówiłaby co innego
    pakiet = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, godz_od, godz_do, trener,"
        " co_ile_tygodni) VALUES (?,'CYKLICZNE-PRZEDSZKOLE','2026-11-03','09:00',"
        "'10:00',?,1)", (lid, "07. Nowak")).lastrowid
    for nr, d in enumerate(["2026-11-03", "2026-11-17", "2026-12-01"], 1):
        conn.execute("INSERT INTO terminy_cyklu (event_id, nr, data) VALUES (?,?,?)",
                     (pakiet, nr, d))

    # 3. DT
    dt_id = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, godz_od, godz_do, trener, ilosc_klas)"
        " VALUES (?,'DT','2026-09-15','09:00','11:00',?,4)", (lid, "05. Kowalski")).lastrowid

    # 4. cykl ODWOŁANY W CAŁOŚCI — ma dojechać do partnera, żeby mógł go wyłączyć
    odwolany = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, godz_od, godz_do, trener, odwolane,"
        " powod_odwolania, odwolal) VALUES (?,'CYKLICZNE','2026-10-06','16:00','17:00',"
        "?,?,?,?)", (lid, "05. Kowalski", "2026-09-01 10:00", "szkoła się rozmyśliła",
                     "01. Sacawa")).lastrowid

    # 5. typ spoza zakresu mostu — nie ma po ich stronie gdzie trafić
    conn.execute("INSERT INTO eventy (lead_id, typ, data, godz_od) "
                 "VALUES (?,'FESTYN','2026-09-20','10:00')", (lid,))
    conn.commit()
    return conn, {"cykl": cykl, "pakiet": pakiet, "dt": dt_id, "odwolany": odwolany}


def po_id(migawka, eid):
    for rec in migawka["cykle"] + migawka["dt"]:
        if rec["source_event_id"] == eid:
            return rec
    return None


# ------------------------------------------------------------------ M1

def m1_zakres(conn, ids):
    print("\nM1 — co wychodzi, a co zostaje")
    m = dane.zbuduj(conn)

    sprawdz("cykle trafiają do sekcji `cykle`", m["liczby"]["cykle"] == 3,
            "%d (reguła + pakiet + odwołany)" % m["liczby"]["cykle"])
    sprawdz("DT trafia do sekcji `dt`", m["liczby"]["dt"] == 1)
    # FESTYN i podobne zostają u nas: po ich stronie nie ma dla nich miejsca,
    # a wysyłanie „na zapas" zamienia most w drugi eksport wszystkiego.
    sprawdz("typ spoza zakresu nie wychodzi wcale",
            all(r["source_type"] != "FESTYN" for r in m["cykle"] + m["dt"]))

    rec = po_id(m, ids["cykl"])
    sprawdz("cykl niesie nazwy pól ODBIORCY, nie nasze",
            all(k in rec for k in ("school", "weekday", "start_time", "end_time",
                                   "starts_on", "region")),
            "school/weekday/start_time/end_time/starts_on/region")
    sprawdz("nasze nazwy pól nie wyciekają do formatu",
            not any(k in rec for k in ("placowka", "godz_od", "cykl_dzien", "trener")))
    sprawdz("id jest stabilnym kluczem zewnętrznym",
            rec["id"] == "leady-v5:event:%d" % ids["cykl"], rec["id"])


# ------------------------------------------------------------------ M2

def m2_zapora(conn, ids):
    print("\nM2 — zapora na dane osobowe (szukamy po WARTOŚCIACH, nie po nazwach pól)")
    m = dane.zbuduj(conn)
    tekst = json.dumps(m, ensure_ascii=False)

    for etykieta, wartosc in (("telefon placówki", TELEFON), ("mail placówki", MAIL),
                              ("osoba kontaktowa", OSOBA),
                              ("notatka handlowca", UWAGI_LEADA),
                              ("notatka przy zajęciach", UWAGI_EVENTU)):
        sprawdz("%s NIE wychodzi w pliku" % etykieta, wartosc not in tekst)

    # Numer telefonu ukryty w wolnym tekście to najczęstsza droga wycieku —
    # dlatego notatek nie wysyłamy w ogóle, a nie „po sprawdzeniu".
    sprawdz("numer z wnętrza notatki też nie wychodzi", "502-111-222" not in tekst)

    xlsx = pliki.sciezka(pliki.PLIK_XLSX)
    most.zapisz(conn)
    with open(xlsx, "rb") as f:
        surowy = f.read()
    # XLSX to zip — teksty siedzą skompresowane, więc czytamy je arkuszem.
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    tresc_xlsx = "\n".join(str(c.value) for ws in wb for row in ws.iter_rows()
                           for c in row if c.value is not None)
    sprawdz("arkusz XLSX też nie niesie telefonu ani maila",
            TELEFON not in tresc_xlsx and MAIL not in tresc_xlsx
            and OSOBA not in tresc_xlsx)
    sprawdz("arkusz powstał i ma trzy zakładki danych",
            {"Cykle", "DT", "Wystąpienia"} <= set(wb.sheetnames), str(wb.sheetnames))
    sprawdz("plik XLSX nie jest pusty", len(surowy) > 4000, "%d bajtów" % len(surowy))


# ------------------------------------------------------------------ M3

def m3_wystapienia(conn, ids):
    print("\nM3 — wystąpienia liczone kalendarzem, nie regułą")
    m = dane.zbuduj(conn)

    rec = po_id(m, ids["cykl"])
    sprawdz("cykl z reguły ma policzone konkretne daty",
            len(rec["occurrences"]) > 4, "%d wystąpień" % len(rec["occurrences"]))
    sprawdz("wszystkie wystąpienia to poniedziałki",
            all(dt.date.fromisoformat(d).isoweekday() == 1 for d in rec["occurrences"]))
    sprawdz("weekday liczony z pierwszego wystąpienia, w ich numeracji 1–7",
            rec["weekday"] == 1, "poniedziałek = 1")

    # To jest sedno: odwołany termin ma zniknąć z grafiku i pojawić się osobno,
    # a nie zniknąć bez śladu — inaczej odbiorca wyśle trenera na nieodbyte zajęcia.
    sprawdz("odwołane wystąpienie NIE jest w occurrences",
            "2026-10-19" not in rec["occurrences"])
    odw = [o["date"] for o in rec["cancelled_occurrences"]]
    sprawdz("odwołane wystąpienie jest w cancelled_occurrences", "2026-10-19" in odw)
    sprawdz("odwołanie niesie powód",
            any(o["reason"] == "ferie w szkole" for o in rec["cancelled_occurrences"]))

    pak = po_id(m, ids["pakiet"])
    sprawdz("pakiet dat wygrywa nad regułą 'co tydzień'",
            pak["occurrences"] == ["2026-11-03", "2026-11-17", "2026-12-01"],
            str(pak["occurrences"]))
    sprawdz("ends_on to ostatni termin pakietu", pak["ends_on"] == "2026-12-01")

    # Cykl z reguły NIE ma daty końca. Ostatnia policzona data to nasz horyzont
    # liczenia — podanie go jako `ends_on` powiedziałoby odbiorcy, że zajęcia
    # się wtedy kończą, a one się nie kończą: przestaliśmy je liczyć.
    sprawdz("cykl bez końca nie udaje, że ma koniec", rec["ends_on"] is None,
            "%d policzonych wystąpień, ends_on=%r"
            % (len(rec["occurrences"]), rec["ends_on"]))
    sprawdz("i mówi wprost, że daty są wyliczone, nie uzgodnione",
            rec["occurrences_agreed"] is False
            and rec["occurrences_horizon_weeks"] == cv.CYKL_HORYZONT_TYGODNI,
            "horyzont %s tyg." % rec["occurrences_horizon_weeks"])
    sprawdz("pakiet odwrotnie — daty uzgodnione, bez horyzontu",
            pak["occurrences_agreed"] is True
            and pak["occurrences_horizon_weeks"] is None)

    # Zgodność z ekranem — ta sama liczba, którą widzi koordynator.
    z_kalendarza = [e["data"] for e in cv.events_for_month(conn, "2026-10")
                    if e["id"] == ids["cykl"]]
    w_pliku = [d for d in rec["occurrences"] if d.startswith("2026-10")]
    sprawdz("plik i kalendarz zgadzają się co do października",
            sorted(z_kalendarza) == sorted(w_pliku),
            "kalendarz %d / plik %d" % (len(z_kalendarza), len(w_pliku)))


# ------------------------------------------------------------------ M4

def m4_odwolania(conn, ids):
    print("\nM4 — odwołany cykl dojeżdża do partnera, żeby mógł go wyłączyć")
    m = dane.zbuduj(conn)
    rec = po_id(m, ids["odwolany"])
    sprawdz("odwołany w całości cykl JEST w pliku", rec is not None)
    sprawdz("i ma stan 'odwolany'", rec and rec["stan"] == "odwolany")
    sprawdz("z powodem odwołania",
            rec and rec.get("cancelled_reason") == "szkoła się rozmyśliła")
    sprawdz("nie ma żadnych żywych wystąpień", rec and rec["occurrences"] == [])
    sprawdz("licznik aktywnych go nie liczy",
            m["liczby"]["cykle_aktywne"] == m["liczby"]["cykle"] - 1,
            "%d aktywnych z %d" % (m["liczby"]["cykle_aktywne"], m["liczby"]["cykle"]))


# ------------------------------------------------------------------ M5

def m5_braki(conn, ids):
    print("\nM5 — braki są WIDOCZNE, nie blokują")
    lid = conn.execute("SELECT id FROM leady LIMIT 1").fetchone()[0]
    # Od poprawek z sierpnia nasza aplikacja pozwala zapisać zajęcia bez godzin —
    # a po ich stronie godzina jest polem wymaganym. Rekord ma dojechać z listą
    # braków, żeby odbiorca odłożył go na bok zamiast wywalić się na NOT NULL.
    ubogi = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data) VALUES (?,'CYKLICZNE','2026-10-07')",
        (lid,)).lastrowid
    conn.commit()
    m = dane.zbuduj(conn)
    rec = po_id(m, ubogi)
    sprawdz("rekord bez godzin mimo wszystko wychodzi", rec is not None)
    sprawdz("z jawną listą braków",
            rec and "start_time" in rec["missing"] and "end_time" in rec["missing"],
            str(rec["missing"]) if rec else "")
    sprawdz("licznik niekompletnych go widzi", m["liczby"]["niekompletne"] >= 1)
    conn.execute("DELETE FROM eventy WHERE id=?", (ubogi,))
    conn.commit()


# ------------------------------------------------------------------ M6

def m6_pliki(conn, ids):
    print("\nM6 — pliki: atomowość, komplet, stan")
    most.zapisz(conn)
    katalog = pliki.KATALOG
    sprawdz("powstał zajecia.json", os.path.exists(pliki.sciezka(pliki.PLIK_DANE)))
    sprawdz("powstał zajecia.xlsx", os.path.exists(pliki.sciezka(pliki.PLIK_XLSX)))
    sprawdz("powstał stan.json", os.path.exists(pliki.sciezka(pliki.PLIK_STAN)))
    sprawdz("nie został żaden plik tymczasowy",
            not [n for n in os.listdir(katalog) if n.startswith(".tmp_")],
            str([n for n in os.listdir(katalog) if n.startswith(".tmp_")]))

    with open(pliki.sciezka(pliki.PLIK_DANE), encoding="utf-8") as f:
        wczytany = json.load(f)
    sprawdz("zapisany plik daje się sparsować w całości",
            wczytany["format"] == dane.WERSJA_FORMATU)
    sprawdz("plik niesie profil bazy (demo nie uda produkcji)",
            wczytany["profil"] == "test", wczytany["profil"])

    stan = pliki.czytaj_stan()
    sprawdz("stan.json mówi, kiedy i ile", bool(stan.get("wygenerowano"))
            and stan["liczby"]["cykle"] == wczytany["liczby"]["cykle"])

    # Nieudany zapis nie ma prawa podmienić dobrego pliku ani zostawić śmiecia.
    stary = open(pliki.sciezka(pliki.PLIK_DANE), encoding="utf-8").read()
    try:
        pliki.zapisz_atomowo(pliki.PLIK_DANE, object())      # celowo zły typ
    except Exception:
        pass
    nowy = open(pliki.sciezka(pliki.PLIK_DANE), encoding="utf-8").read()
    sprawdz("nieudany zapis zostawia stary plik nietknięty", stary == nowy)
    sprawdz("i nie zostawia po sobie śmiecia",
            not [n for n in os.listdir(katalog) if n.startswith(".tmp_")])


# ------------------------------------------------------------------ M7

def m7_zmiany(conn, ids):
    print("\nM7 — dziennik zmian i dławik")
    most.zapisz(conn)
    dziennik = pliki.sciezka(pliki.PLIK_ZMIANY)
    przed = os.path.getsize(dziennik) if os.path.exists(dziennik) else 0

    # nic nie zmieniamy → nie ma czego dopisać
    most.zapisz(conn)
    po_pustym = os.path.getsize(dziennik) if os.path.exists(dziennik) else 0
    sprawdz("przebieg bez zmian nie dopisuje nic do dziennika", przed == po_pustym)

    lid = conn.execute("SELECT id FROM leady LIMIT 1").fetchone()[0]
    nowy = conn.execute(
        "INSERT INTO eventy (lead_id, typ, data, godz_od, godz_do, trener) "
        "VALUES (?,'CYKLICZNE','2026-10-08','12:00','13:00','07. Nowak')",
        (lid,)).lastrowid
    conn.commit()
    most.zapisz(conn)
    wpisy = [json.loads(w) for w in open(dziennik, encoding="utf-8") if w.strip()]
    sprawdz("nowy cykl daje wpis 'dodane'",
            any(w["co"] == "dodane" and w["id"].endswith(":%d" % nowy) for w in wpisy))

    conn.execute("UPDATE eventy SET godz_do='14:00' WHERE id=?", (nowy,))
    conn.commit()
    most.zapisz(conn)
    wpisy = [json.loads(w) for w in open(dziennik, encoding="utf-8") if w.strip()]
    sprawdz("zmiana godziny daje wpis 'zmienione'",
            any(w["co"] == "zmienione" and w["id"].endswith(":%d" % nowy) for w in wpisy))

    conn.execute("DELETE FROM eventy WHERE id=?", (nowy,))
    conn.commit()
    most.zapisz(conn)
    wpisy = [json.loads(w) for w in open(dziennik, encoding="utf-8") if w.strip()]
    # Skasowane u nas ma u odbiorcy zostać WYŁĄCZONE, nie skasowane — przy jego
    # zajęciach mogą już wisieć zastępstwa i wpisy rozliczeniowe.
    sprawdz("skasowany cykl daje wpis 'zniknelo'",
            any(w["co"] == "zniknelo" and w["id"].endswith(":%d" % nowy) for w in wpisy))


# ------------------------------------------------------------------ M8

def m8_zegar(conn, ids):
    print("\nM8 — dwa zegary: zmiana i bicie serca")
    teraz = dt.datetime(2026, 9, 1, 12, 0, 0)
    most.zapisz(conn, teraz=teraz)

    pora, powod = most.czy_pora(conn, teraz=teraz + dt.timedelta(seconds=5))
    sprawdz("zaraz po zapisie nie ma pory", not pora, powod)

    db.meta_set(conn, db.MOST_BRUDNY, "1")
    conn.commit()
    pora, _ = most.czy_pora(conn, teraz=teraz + dt.timedelta(seconds=5))
    sprawdz("zmiana w bazie NIE przebija minimalnego odstępu", not pora,
            "chroni przed przepisywaniem pliku kilka razy w trakcie jednego zapisu")

    pora, powod = most.czy_pora(conn, teraz=teraz + dt.timedelta(seconds=30))
    sprawdz("po odstępie zmiana wymusza zapis", pora, powod)

    db.meta_set(conn, db.MOST_BRUDNY, "")
    conn.commit()
    pora, _ = most.czy_pora(conn, teraz=teraz + dt.timedelta(minutes=5))
    sprawdz("bez zmian 5 minut później nadal nie ma pory", not pora)
    pora, powod = most.czy_pora(conn, teraz=teraz + dt.timedelta(minutes=61))
    sprawdz("po godzinie ciszy most odświeża się sam", pora, powod)

    # Znacznik stawia `zapisz_log`, czyli jedyne miejsce, przez które przechodzi
    # każdy zapis eventu i leada.
    db.meta_set(conn, db.MOST_BRUDNY, "")
    conn.commit()
    lid = conn.execute("SELECT id FROM leady LIMIT 1").fetchone()[0]
    db.zapisz_log(conn, lead_id=lid, kto="test", co="próba")
    conn.commit()
    sprawdz("zapis do logu sam oznacza bazę jako brudną",
            db.meta_get(conn, db.MOST_BRUDNY) == "1")


# ------------------------------------------------------------------ M9

def m9_odpornosc(conn, ids):
    print("\nM9 — most jest dodatkiem i nie ma prawa wywrócić żądania")
    stary = pliki.KATALOG
    try:
        # katalog, którego nie da się utworzyć — symuluje odmontowany wolumen
        pliki.KATALOG = os.path.join(stary, "plik-nie-katalog", "most")
        open(os.path.join(stary, "plik-nie-katalog"), "w").close()
        wynik = most.przeglad(conn, teraz=dt.datetime(2027, 1, 1, 12, 0))
        sprawdz("awaria katalogu nie rzuca wyjątkiem", wynik is None)
    finally:
        pliki.KATALOG = stary

    sprawdz("po awarii most wraca do pracy",
            most.przeglad(conn, teraz=dt.datetime(2027, 1, 1, 13, 0)) is not None)


def main():
    conn, ids = zbuduj_baze()
    try:
        m1_zakres(conn, ids)
        m2_zapora(conn, ids)
        m3_wystapienia(conn, ids)
        m4_odwolania(conn, ids)
        m5_braki(conn, ids)
        m6_pliki(conn, ids)
        m7_zmiany(conn, ids)
        m8_zegar(conn, ids)
        m9_odpornosc(conn, ids)
    finally:
        conn.close()

    ile_ok = sum(1 for _n, w in WYNIKI if w)
    print("\n%d/%d" % (ile_ok, len(WYNIKI)))
    zle = [n for n, w in WYNIKI if not w]
    if zle:
        print("BLEDY: " + " · ".join(zle))
    return 0 if not zle else 1


if __name__ == "__main__":
    raise SystemExit(main())
