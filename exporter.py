# -*- coding: utf-8 -*-
"""
Eksport do XLSX — w układzie, który klient zna z własnego arkusza.

Dwie zasady:

1. EKSPORTUJEMY DOKŁADNIE TO, CO WIDAĆ. Wiersze przychodzą z `repo.filtruj_leady`,
   czyli z tej samej funkcji, która zasila ekran. Jeśli koordynator wyfiltruje
   Katowice + Chytry + „po terminie", to w pliku będzie Katowice + Chytry + po terminie.
   To było wprost zgłoszone życzenie („gdy wyfiltruje wartości, chcę mieć możliwość
   pobrania wyfiltrowanego do excela").

2. UKŁAD KOLUMN JAK U NICH — nagłówki i kolejność z ich arkusza (`Handlowiec`,
   `Status szkoły`, …, `Librus WYPEŁNIA JULIA`). Dzięki temu plik da się wkleić
   w istniejący proces i nikt nie musi się przestawiać.

Dodatkowo dokładamy trzy arkusze, których w ich pliku nie było, a które robią różnicę:
`Spotkania` (1 wiersz = 1 spotkanie — dowód, że dubli nie gubimy), `Kolizje`
i `Filtr` (co dokładnie było ustawione przy eksporcie).
"""
import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import calendar_view as cv
from db import LEAD_FIELDS, JULIA_FIELDS

# Kolumny arkusza „Leady" — (nagłówek dla klienta, klucz w wierszu, szerokość)
KOLUMNY = [
    ("Handlowiec",                "handlowiec",         18),
    ("Status szkoły",             "status_szkoly",       16),
    ("Status realizacji",         "status_realizacji",   30),
    ("death line",                "deadline",            12),
    ("Miejscowość",               "miejscowosc",         20),
    ("Numer placówki",            "placowka",            42),
    ("Adres placówki",            "adres",               28),
    ("Osoby decyzyjne i kontakt", "osoba_kontakt",       26),
    ("numer telefonu",            "telefon",             16),
    ("mail",                      "mail",                28),
    ("Uwagi",                     "uwagi",               34),
    ("Do zrobienia",              "do_zrobienia",        30),
    ("DT",                        "dt",                  14),
    ("Data DT",                   "dt_data",             12),
    ("Godzina DT od",             "dt_godz_od",          12),
    ("Godzina DT do",             "dt_godz_do",          12),
    ("Prowadzący DT",             "dt_trener",           20),
    ("Numer sali DT",             "dt_sala",             12),
    ("mail propozycja lub ustalenie DT", "mail_dt",      24),
    ("Ilość klas 1-4",            "dt_klas",              9),
    ("Ilość dzieci w klasach",    "dt_dzieci",           10),
    ("Mail do rodziców na dziennik elektroniczny", "mail_rodzice", 16),
    ("Cykle",                     "cykle",               10),
    ("Liczba DT",                 "n_dt",                 9),
    ("Liczba zajęć cyklicznych",  "n_cykl",               9),
    ("Typ placówki",              "typ_placowki",        26),
    ("Nr RSPO",                   "rspo",                12),
    ("Mail z wnioskiem o wynajem sali", "mail_wynajem",  16),
] + [(etykieta, klucz, 16) for etykieta, klucz, _typ, naglowek in JULIA_FIELDS]

KOLUMNY_EVENTY = [
    ("Typ",            "typ",             12),
    ("Data",           "data",            12),
    ("Dzień",          "_dow",            12),
    ("Godz. od",       "godz_od",         10),
    ("Godz. do",       "godz_do",         10),
    ("Prowadzący",     "trener",          20),
    ("Drugi prowadzący", "trener2",       20),
    ("Zastępstwo",     "zastepstwo",      18),
    ("Drukarz",        "drukarz",         18),
    ("Placówka",       "placowka",        40),
    ("Miejscowość",    "miejscowosc",     18),
    ("Nr sali",        "numer_sali",      10),
    ("Grupa",          "grupa",            8),
    ("Sprzęt",         "sprzet",          22),
    ("Ilość klas",     "ilosc_klas",       9),
    ("Ilość dzieci",   "ilosc_dzieci",    10),
    ("Dzień cyklu",    "cykl_dzien",      14),
    ("Co ile tygodni", "co_ile_tygodni",   9),
    ("Kod Tinkercad",  "kod_tinkercad",   16),
    ("Link Tinkercad", "link_tinkercad",  38),
    ("Kolizja",        "_kolizja",         9),
    ("Handlowiec",     "handlowiec",      18),
]

ZIELONY = PatternFill("solid", fgColor="93C47D")     # ich kolor nagłówków
ROZOWY = PatternFill("solid", fgColor="F4CCCC")      # ich kolor kolumn Julki
MAGENTA = PatternFill("solid", fgColor="FFF1F4")
BOLD = Font(bold=True)
CIENKA = Border(bottom=Side(style="thin", color="D7D3D3"))

ETYKIETY_ZAKRESU = {
    "": "wszystkie",
    "nieprzydzielone": "do rozdania (nieprzydzielone)",
    "przydzielone": "u handlowców",
    "umowione": "DT umówione",
    "niewykorzystane": "niewykorzystane rekordy",
    "po_terminie": "po terminie",
    "cykle": "z zajęciami cyklicznymi",
    "pin": "plan tygodnia (przypięte)",
}


def _naglowki(ws, kolumny, julia_od=None):
    for i, (etykieta, _klucz, szer) in enumerate(kolumny, 1):
        c = ws.cell(row=1, column=i, value=etykieta)
        c.font = BOLD
        c.fill = ROZOWY if (julia_od and i >= julia_od) else ZIELONY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(i)].width = szer
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(kolumny))


def _wartosc(v):
    """openpyxl nie zapisze dowolnego obiektu — sprowadzamy do tekstu/liczby."""
    if v is None or isinstance(v, (str, int, float, dt.date, dt.datetime)):
        return v
    return str(v)


def build_workbook(conn, rows, f=None, month=None):
    """
    rows — wiersze z `repo.filtruj_leady` (już przefiltrowane!)
    f    — słownik filtrów, żeby zapisać w arkuszu „Filtr", co było ustawione
    """
    f = f or {}
    wb = Workbook()

    # ---------------------------------------------------------- Leady
    ws = wb.active
    ws.title = "Leady"
    julia_od = len(KOLUMNY) - len(JULIA_FIELDS) + 1
    _naglowki(ws, KOLUMNY, julia_od=julia_od)
    dzis = dt.date.today().isoformat()
    for r_i, row in enumerate(rows, 2):
        for c_i, (_etykieta, klucz, _szer) in enumerate(KOLUMNY, 1):
            cell = ws.cell(row=r_i, column=c_i, value=_wartosc(row.get(klucz)))
            cell.border = CIENKA
            cell.alignment = Alignment(vertical="top", wrap_text=(klucz in
                                       ("uwagi", "do_zrobienia", "placowka")))
            if klucz == "deadline" and row.get("po_terminie"):
                cell.fill = MAGENTA
                cell.font = Font(bold=True, color="AA0B56")

    # ---------------------------------------------------------- Spotkania
    ws2 = wb.create_sheet("Spotkania")
    _naglowki(ws2, KOLUMNY_EVENTY)
    lead_ids = {row["id"] for row in rows}
    miesiace = cv.available_months(conn)
    evs = []
    for mm in ([month] if month else miesiace):
        evs += cv.events_for_month(conn, mm)
    kolizje = cv.find_collisions(evs)
    evs = [e for e in evs if e["lead_id"] in lead_ids]
    evs.sort(key=lambda e: (e["data"], e["godz_od"] or "99:99", e["trener"] or ""))
    for r_i, e in enumerate(evs, 2):
        try:
            d = dt.date.fromisoformat(e["data"])
            e["_dow"] = cv.DNI[d.weekday()]
        except (ValueError, TypeError):
            e["_dow"] = ""
        e["_kolizja"] = "TAK" if e["_key"] in kolizje else ""
        for c_i, (_etykieta, klucz, _szer) in enumerate(KOLUMNY_EVENTY, 1):
            cell = ws2.cell(row=r_i, column=c_i, value=_wartosc(e.get(klucz)))
            cell.border = CIENKA
            if e["_kolizja"]:
                cell.fill = MAGENTA

    # ---------------------------------------------------------- Kolizje
    ws3 = wb.create_sheet("Kolizje")
    _naglowki(ws3, [("Data", "data", 12), ("Dzień", "_dow", 12),
                    ("Prowadzący", "trener", 20), ("Godz. od", "godz_od", 10),
                    ("Godz. do", "godz_do", 10), ("Placówka", "placowka", 40),
                    ("Miejscowość", "miejscowosc", 18), ("Typ", "typ", 12)])
    kol = [e for e in evs if e["_kolizja"]]
    for r_i, e in enumerate(kol, 2):
        for c_i, klucz in enumerate(["data", "_dow", "trener", "godz_od", "godz_do",
                                     "placowka", "miejscowosc", "typ"], 1):
            ws3.cell(row=r_i, column=c_i, value=_wartosc(e.get(klucz))).border = CIENKA
    if not kol:
        ws3.cell(row=2, column=1, value="Brak kolizji w danych.")

    # ---------------------------------------------------------- Filtr
    ws4 = wb.create_sheet("Filtr")
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 46
    ws4.cell(row=1, column=1, value="Eksport z Systemu Leadów").font = Font(bold=True, size=13)
    opis = [
        ("data eksportu", dzis),
        ("zakres", ETYKIETY_ZAKRESU.get(f.get("zakres", ""), f.get("zakres") or "wszystkie")),
        ("handlowiec", f.get("handlowiec") or "— wszyscy —"),
        ("miejscowość", f.get("miasto") or "— wszystkie —"),
        ("status realizacji", f.get("status") or "— wszystkie —"),
        ("status szkoły", f.get("status_szkoly") or "— wszystkie —"),
        ("typ placówki", f.get("typ") or "— wszystkie —"),
        ("DT", f.get("dt") or "— wszystkie —"),
        ("szukana fraza", f.get("q") or "—"),
        ("", ""),
        ("wyeksportowanych leadów", len(rows)),
        ("wyeksportowanych spotkań", len(evs)),
        ("w tym kolizji trenera", len(kol)),
    ]
    for i, (k, v) in enumerate(opis, 3):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=bool(k))
        ws4.cell(row=i, column=2, value=_wartosc(v))
    ws4.cell(row=len(opis) + 4, column=1,
             value="Arkusz „Spotkania” to jeden wiersz na jedno spotkanie "
                   "— dlatego trener z dwoma DT w jednym dniu ma tu dwa wiersze, "
                   "a nie jeden.")
    ws4.cell(row=len(opis) + 4, column=1).alignment = Alignment(wrap_text=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
